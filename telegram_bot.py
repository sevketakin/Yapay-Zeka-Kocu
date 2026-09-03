"""
Koçum - Telegram Bot Sürümü
================================================================

Chainlit sürümüyle aynı beyni (Gemini + video/eski sohbet arşivi)
kullanır, ama Telegram üzerinden çalışır. Basit, kendi kurduğumuz
bir veritabanı tablosuyla (Chainlit'in karmaşık şemasına gerek
kalmadan) her kullanıcının kendi sohbet geçmişini ayrı ayrı tutar.

Birden fazla kişi (örn. sen ve kız arkadaşın) aynı botu kullanabilir,
Telegram otomatik olarak kim yazdığını ayırt eder, herkesin kendi
geçmişi ayrı kalır. Ortak olan tek şey video/bilgi arşivi.

Gereksinim:
    pip install python-telegram-bot chromadb google-genai openpyxl psycopg2-binary requests

Ortam değişkenleri (Railway'de ayarlanacak):
    TELEGRAM_BOT_TOKEN, GEMINI_API_ANAHTARI, DATABASE_URL, VERI_KLASORU
"""

import os
import re
import asyncio
import anthropic
import json
import uuid
import zipfile
import glob
import base64
from io import BytesIO
from datetime import datetime, timedelta

from google import genai
import chromadb
from chromadb import Documents, EmbeddingFunction, Embeddings
import psycopg2
import requests

from telegram import Update, InputFile, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    Application, MessageHandler, CommandHandler, CallbackQueryHandler,
    ContextTypes, filters,
)

# ============== AYARLAR ==============
VERI_KLASORU = os.environ.get("VERI_KLASORU", ".")
DB_KLASORU = os.path.join(VERI_KLASORU, "veritabani")
GEMINI_API_ANAHTARI = os.environ.get("GEMINI_API_ANAHTARI", "")
TELEGRAM_BOT_TOKEN = os.environ.get("TELEGRAM_BOT_TOKEN", "")
DATABASE_URL = os.environ.get("DATABASE_URL", "")
KOLEKSIYON_ADI = "video_transkriptleri"
GEMINI_MODEL_LISTESI = [
    "gemini-flash-latest",
    "gemini-3.5-flash",
    "gemini-3.1-flash-lite",
    "gemini-3-flash-preview",
]

# SADECE ana koçluk cevabı (cevap_uret) için kullanılır — Pro'yu önce
# dener (daha karmaşık kurallara daha sadık kalıyor), olmazsa/hata
# verirse otomatik Flash'e düşer. Diğer tüm işler (takvim/excel/profil/
# yemek/ses) hâlâ GEMINI_MODEL_LISTESI'ni (ucuz Flash) kullanmaya devam
# ediyor — maliyet sadece en kritik yerde artıyor.
# NOT: 'gemini-3.1-pro' (preview'sız) VE 'gemini-2.5-pro' test edildi,
# ikisi de kalıcı olarak 404 veriyor (2.5-pro artık yeni kullanıcılara
# hiç açılmıyor) — listede YOK, gereksiz başarısız deneme/gecikme
# yaratmasınlar diye. Doğrulanan, çalışan tek Pro modeli en başta.
ANA_CEVAP_MODEL_LISTESI = [
    "gemini-3.1-pro-preview",
] + GEMINI_MODEL_LISTESI

# Claude entegrasyonu — ana cevap için ÖNCE Claude denenir (Opus -> Sonnet),
# ikisi de başarısız olursa (Anthropic tamamen çökse bile) yukarıdaki
# Gemini zincirine (Pro -> Flash) güvenle düşer. ANTHROPIC_API_KEY
# tanımlı değilse Claude denemesi tamamen atlanır, direkt Gemini kullanılır.
ANTHROPIC_API_KEY = os.environ.get("ANTHROPIC_API_KEY", "")
CLAUDE_MODEL_LISTESI = ["claude-sonnet-5", "claude-opus-4-8"]

KAC_PARCA_GETIRILSIN = 16
UYGULAMA_ADI = "Koçum"

STRAVA_CLIENT_ID = os.environ.get("STRAVA_CLIENT_ID", "")
STRAVA_CLIENT_SECRET = os.environ.get("STRAVA_CLIENT_SECRET", "")

# Buraya, "yumuşak ton" ile konuşulmasını istediğin kullanıcıların
# Telegram ID'lerini ekleyebilirsin (örn. kız arkadaşının ID'si).
# ID'yi öğrenmek için: bota /id yazması yeterli, sana ID'sini gösterir.
YUMUSAK_TON_KULLANICILARI = set()

GITHUB_ZIP_URL = (
    "https://github.com/sevketakin/Yapay-Zeka-Kocu/releases/download/v1/veritabani.zip"
)

AYLAR_TR = ["Ocak", "Şubat", "Mart", "Nisan", "Mayıs", "Haziran",
            "Temmuz", "Ağustos", "Eylül", "Ekim", "Kasım", "Aralık"]
GUNLER_TR = ["Pazartesi", "Salı", "Çarşamba", "Perşembe", "Cuma", "Cumartesi", "Pazar"]
GUN_INDEX = {"Pazartesi": 0, "Salı": 1, "Çarşamba": 2, "Perşembe": 3,
             "Cuma": 4, "Cumartesi": 5, "Pazar": 6}


# ============== VERİTABANI ZIP'İNİ OTOMATİK İNDİRME + AÇMA ==============
def _zip_saglam_mi(yol):
    try:
        with zipfile.ZipFile(yol, 'r') as z:
            return z.testzip() is None
    except Exception:
        return False


def _veritabanini_hazirla():
    _tamamlandi_isareti = os.path.join(VERI_KLASORU, "_veritabani_tamamlandi.txt")
    if os.path.exists(_tamamlandi_isareti):
        return

    if os.path.exists(DB_KLASORU):
        import shutil as _shutil
        print("Yarım kalmış eski veritabanı klasörü bulundu, temizleniyor...")
        _shutil.rmtree(DB_KLASORU, ignore_errors=True)

    os.makedirs(VERI_KLASORU, exist_ok=True)
    _tek_zip = os.path.join(VERI_KLASORU, "veritabani.zip")

    if os.path.exists(_tek_zip) and not _zip_saglam_mi(_tek_zip):
        os.remove(_tek_zip)

    if not os.path.exists(_tek_zip):
        try:
            print("Veritabanı GitHub Release'ten indiriliyor... (bu birkaç dakika sürebilir)")
            with requests.get(GITHUB_ZIP_URL, stream=True, allow_redirects=True, timeout=120) as yanit:
                yanit.raise_for_status()
                with open(_tek_zip, "wb") as f:
                    for parca in yanit.iter_content(chunk_size=8 * 1024 * 1024):
                        f.write(parca)
            print("İndirme tamamlandı.")
        except Exception as e:
            print(f"GitHub'dan indirme başarısız: {e}")
            if os.path.exists(_tek_zip):
                os.remove(_tek_zip)
            return

    if not _zip_saglam_mi(_tek_zip):
        print("HATA: İndirilen zip bozuk.")
        return

    print(f"'{_tek_zip}' açılıyor... (bu biraz sürebilir)")
    with zipfile.ZipFile(_tek_zip, 'r') as z:
        z.extractall(VERI_KLASORU)
    os.remove(_tek_zip)

    with open(_tamamlandi_isareti, "w") as f:
        f.write("ok")
    print("Veritabanı hazır.")


# ============== GEMINI EMBEDDING (arama için) ==============
class GeminiEmbeddingFonksiyonu(EmbeddingFunction):
    def __init__(self, api_anahtari):
        self.client = genai.Client(api_key=api_anahtari)

    def __call__(self, input: Documents) -> Embeddings:
        return [self._tek_embedding_al(metin) for metin in input]

    def _tek_embedding_al(self, metin, max_deneme=3):
        import time
        for deneme in range(1, max_deneme + 1):
            try:
                sonuc = self.client.models.embed_content(
                    model="gemini-embedding-001", contents=metin,
                )
                return sonuc.embeddings[0].values
            except Exception as e:
                hata_metni = str(e)
                if "429" in hata_metni or "RESOURCE_EXHAUSTED" in hata_metni or "UNAVAILABLE" in hata_metni:
                    time.sleep(5)
                    continue
                raise
        raise RuntimeError(f"Embedding alınamadı: {metin[:50]}...")


_client_gemini = None
_koleksiyon = None


def istemcileri_al():
    global _client_gemini, _koleksiyon
    if _client_gemini is None:
        _client_gemini = genai.Client(api_key=GEMINI_API_ANAHTARI)
    if _koleksiyon is None:
        embed_fonksiyonu = GeminiEmbeddingFonksiyonu(GEMINI_API_ANAHTARI)
        db = chromadb.PersistentClient(path=DB_KLASORU)
        _koleksiyon = db.get_collection(name=KOLEKSIYON_ADI, embedding_function=embed_fonksiyonu)
    return _client_gemini, _koleksiyon


def baglami_hazirla(bulunan_parcalar):
    metinler = bulunan_parcalar['documents'][0]
    metadatalar = bulunan_parcalar['metadatas'][0]
    baglam_parcalari = []
    kaynaklar = []
    for metin, meta in zip(metinler, metadatalar):
        video_id = meta.get('video_id', 'bilinmiyor')
        kaynak_turu = meta.get('kaynak', '')
        onizleme = metin[:250].strip() + ("..." if len(metin) > 250 else "")
        if kaynak_turu == "eski_sohbet":
            baglam_parcalari.append(
                f"[GERÇEK KİŞİSEL GEÇMİŞ — {video_id} — bu kullanıcının SENİNLE "
                f"gerçekten daha önce konuştuğu bir şey, genel bir video örneği DEĞİL]\n{metin}"
            )
            kaynaklar.append({"link": None, "baslik": video_id, "onizleme": onizleme})
        else:
            link = f"https://youtube.com/watch?v={video_id}"
            baglam_parcalari.append(
                f"[Genel video içeriği (BAŞKA insanların/koçların anlattığı örnekler, "
                f"kullanıcının kendi kişisel geçmişi DEĞİL): {link}]\n{metin}"
            )
            kaynaklar.append({"link": link, "baslik": None, "onizleme": onizleme})
    return "\n\n---\n\n".join(baglam_parcalari), kaynaklar


# ============== BASİT SOHBET GEÇMİŞİ (kendi tablomuz) ==============
def _basit_semayi_hazirla():
    if not DATABASE_URL:
        return
    try:
        baglanti = psycopg2.connect(DATABASE_URL)
        baglanti.autocommit = True
        imlec = baglanti.cursor()
        imlec.execute("""
            CREATE TABLE IF NOT EXISTS tg_mesajlar (
                id SERIAL PRIMARY KEY,
                kullanici_id BIGINT NOT NULL,
                rol TEXT NOT NULL,
                icerik TEXT NOT NULL,
                zaman TIMESTAMP DEFAULT NOW()
            );
            CREATE TABLE IF NOT EXISTS tg_ayarlar (
                kullanici_id BIGINT PRIMARY KEY,
                yumusak_ton BOOLEAN DEFAULT FALSE
            );
            CREATE TABLE IF NOT EXISTS strava_baglantilar (
                kullanici_id BIGINT PRIMARY KEY,
                refresh_token TEXT NOT NULL,
                athlete_id BIGINT,
                son_gorulen_aktivite_id BIGINT DEFAULT 0
            );
            CREATE TABLE IF NOT EXISTS kullanici_profil (
                kullanici_id BIGINT PRIMARY KEY,
                profil TEXT DEFAULT ''
            );
            CREATE TABLE IF NOT EXISTS vucut_olculeri (
                id SERIAL PRIMARY KEY,
                kullanici_id BIGINT NOT NULL,
                tarih TIMESTAMP DEFAULT NOW(),
                kilo TEXT,
                kol TEXT,
                bel TEXT,
                kalca TEXT,
                gogus TEXT,
                bacak TEXT
            );
            ALTER TABLE vucut_olculeri ADD COLUMN IF NOT EXISTS omuz TEXT;
            ALTER TABLE tg_ayarlar ADD COLUMN IF NOT EXISTS sabah_mesaji BOOLEAN DEFAULT FALSE;
            ALTER TABLE tg_ayarlar ADD COLUMN IF NOT EXISTS olcum_hatirlatma BOOLEAN DEFAULT TRUE;
            ALTER TABLE tg_ayarlar ADD COLUMN IF NOT EXISTS sesli_cevap BOOLEAN DEFAULT FALSE;
            ALTER TABLE tg_ayarlar ADD COLUMN IF NOT EXISTS son_sabah_mesaji_tarihi DATE;
            CREATE TABLE IF NOT EXISTS beslenme_kayitlari (
                id SERIAL PRIMARY KEY,
                kullanici_id BIGINT NOT NULL,
                tarih TIMESTAMP DEFAULT NOW(),
                aciklama TEXT,
                tahmini_kalori INTEGER,
                tahmini_protein INTEGER,
                tahmini_karbonhidrat INTEGER,
                tahmini_yag INTEGER
            );
            CREATE TABLE IF NOT EXISTS intervals_baglantilar (
                kullanici_id BIGINT PRIMARY KEY,
                api_key TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS intervals_bildirilen_aktiviteler (
                kullanici_id BIGINT NOT NULL,
                aktivite_id TEXT NOT NULL,
                PRIMARY KEY (kullanici_id, aktivite_id)
            );
            CREATE TABLE IF NOT EXISTS antrenman_gunlugu (
                id SERIAL PRIMARY KEY,
                kullanici_id BIGINT NOT NULL,
                tarih DATE DEFAULT CURRENT_DATE,
                aciklama TEXT NOT NULL
            );
        """)
        imlec.close()
        baglanti.close()
        print("Telegram bot tabloları hazır.")
    except Exception as e:
        print(f"Tablo hazırlanırken hata: {e}")


def _gecmis_tarih_etiketi(zaman):
    """Bir geçmiş mesajın, BUGÜNE göre kaç gün önce olduğunu Türkçe
    etiketler — böylece model 'dün konuştuğumuz şey' ile 'bugün
    konuştuğumuz şey'i birbirinden ayırt edebilir."""
    try:
        from zoneinfo import ZoneInfo
        if zaman.tzinfo is None:
            zaman_tr = zaman.replace(tzinfo=ZoneInfo("UTC")).astimezone(ZoneInfo("Europe/Istanbul"))
        else:
            zaman_tr = zaman.astimezone(ZoneInfo("Europe/Istanbul"))
    except Exception:
        zaman_tr = zaman

    bugun = _turkiye_simdi().date()
    fark = (bugun - zaman_tr.date()).days
    saat_str = zaman_tr.strftime("%H:%M")

    if fark <= 0:
        return f"BUGÜN, saat {saat_str}"
    elif fark == 1:
        return f"DÜN, saat {saat_str}"
    else:
        return f"{fark} gün önce ({zaman_tr.day} {AYLAR_TR[zaman_tr.month - 1]}), saat {saat_str}"


def gecmisi_oku(kullanici_id, limit=40):
    if not DATABASE_URL:
        return []
    try:
        baglanti = psycopg2.connect(DATABASE_URL)
        imlec = baglanti.cursor()
        imlec.execute(
            "SELECT rol, icerik, zaman FROM ("
            "  SELECT rol, icerik, zaman FROM tg_mesajlar "
            "  WHERE kullanici_id = %s ORDER BY zaman DESC LIMIT %s"
            ") alt ORDER BY zaman ASC",
            (kullanici_id, limit),
        )
        satirlar = imlec.fetchall()
        imlec.close()
        baglanti.close()
        sonuc = []
        for rol, icerik, zaman in satirlar:
            etiket = _gecmis_tarih_etiketi(zaman)
            icerik_etiketli = f"[{etiket}] {icerik}"
            sonuc.append({"role": rol, "parts": [{"text": icerik_etiketli}]})
        return sonuc
    except Exception as e:
        print(f"Geçmiş okunurken hata: {e}")
        return []


def mesaji_kaydet(kullanici_id, rol, icerik):
    if not DATABASE_URL:
        return
    try:
        baglanti = psycopg2.connect(DATABASE_URL)
        baglanti.autocommit = True
        imlec = baglanti.cursor()
        imlec.execute(
            "INSERT INTO tg_mesajlar (kullanici_id, rol, icerik) VALUES (%s, %s, %s)",
            (kullanici_id, rol, icerik),
        )
        imlec.close()
        baglanti.close()
    except Exception as e:
        print(f"Mesaj kaydedilirken hata: {e}")


def yumusak_ton_mu(kullanici_id):
    if kullanici_id in YUMUSAK_TON_KULLANICILARI:
        return True
    if not DATABASE_URL:
        return False
    try:
        baglanti = psycopg2.connect(DATABASE_URL)
        imlec = baglanti.cursor()
        imlec.execute("SELECT yumusak_ton FROM tg_ayarlar WHERE kullanici_id = %s", (kullanici_id,))
        satir = imlec.fetchone()
        imlec.close()
        baglanti.close()
        return bool(satir[0]) if satir else False
    except Exception:
        return False


def yumusak_ton_ayarla(kullanici_id, deger):
    if not DATABASE_URL:
        return
    try:
        baglanti = psycopg2.connect(DATABASE_URL)
        baglanti.autocommit = True
        imlec = baglanti.cursor()
        imlec.execute(
            "INSERT INTO tg_ayarlar (kullanici_id, yumusak_ton) VALUES (%s, %s) "
            "ON CONFLICT (kullanici_id) DO UPDATE SET yumusak_ton = %s",
            (kullanici_id, deger, deger),
        )
        imlec.close()
        baglanti.close()
    except Exception as e:
        print(f"Ayar kaydedilirken hata: {e}")


# ============== STRAVA ENTEGRASYONU ==============
def strava_baglantisini_kaydet(kullanici_id, refresh_token, athlete_id=None):
    if not DATABASE_URL:
        return
    baglanti = psycopg2.connect(DATABASE_URL)
    baglanti.autocommit = True
    imlec = baglanti.cursor()
    imlec.execute(
        "INSERT INTO strava_baglantilar (kullanici_id, refresh_token, athlete_id) "
        "VALUES (%s, %s, %s) "
        "ON CONFLICT (kullanici_id) DO UPDATE SET refresh_token = %s, athlete_id = %s",
        (kullanici_id, refresh_token, athlete_id, refresh_token, athlete_id),
    )
    imlec.close()
    baglanti.close()


def strava_baglantisini_getir(kullanici_id):
    if not DATABASE_URL:
        return None
    baglanti = psycopg2.connect(DATABASE_URL)
    imlec = baglanti.cursor()
    imlec.execute(
        "SELECT refresh_token, athlete_id, son_gorulen_aktivite_id "
        "FROM strava_baglantilar WHERE kullanici_id = %s",
        (kullanici_id,),
    )
    satir = imlec.fetchone()
    imlec.close()
    baglanti.close()
    if not satir:
        return None
    return {"refresh_token": satir[0], "athlete_id": satir[1], "son_gorulen": satir[2]}


def strava_son_gorulen_guncelle(kullanici_id, aktivite_id):
    if not DATABASE_URL:
        return
    baglanti = psycopg2.connect(DATABASE_URL)
    baglanti.autocommit = True
    imlec = baglanti.cursor()
    imlec.execute(
        "UPDATE strava_baglantilar SET son_gorulen_aktivite_id = %s WHERE kullanici_id = %s",
        (aktivite_id, kullanici_id),
    )
    imlec.close()
    baglanti.close()


def strava_erisim_tokeni_al(refresh_token):
    """refresh_token'dan (kalıcı) taze bir access_token (6 saatlik) üretir."""
    yanit = requests.post(
        "https://www.strava.com/oauth/token",
        data={
            "client_id": STRAVA_CLIENT_ID,
            "client_secret": STRAVA_CLIENT_SECRET,
            "refresh_token": refresh_token,
            "grant_type": "refresh_token",
        },
        timeout=15,
    )
    yanit.raise_for_status()
    return yanit.json()["access_token"]


def strava_son_aktiviteleri_getir(access_token, kac_tane=5):
    yanit = requests.get(
        "https://www.strava.com/api/v3/athlete/activities",
        headers={"Authorization": f"Bearer {access_token}"},
        params={"per_page": kac_tane},
        timeout=15,
    )
    yanit.raise_for_status()
    return yanit.json()


def strava_aktiviteyi_metne_cevir(aktivite):
    isim = aktivite.get("name", "Aktivite")
    tur = aktivite.get("type", "")
    mesafe_km = (aktivite.get("distance", 0) or 0) / 1000
    sure_dk = (aktivite.get("moving_time", 0) or 0) / 60
    ort_nabiz = aktivite.get("average_heartrate")
    yukseklik = aktivite.get("total_elevation_gain")
    tarih = aktivite.get("start_date_local", "")

    satirlar = [f"Aktivite: {isim} ({tur})", f"Tarih: {tarih}", f"Süre: {sure_dk:.0f} dakika"]
    if mesafe_km > 0:
        satirlar.append(f"Mesafe: {mesafe_km:.2f} km")
    if ort_nabiz:
        satirlar.append(f"Ortalama nabız: {ort_nabiz:.0f}")
    if yukseklik:
        satirlar.append(f"Toplam tırmanış: {yukseklik:.0f} m")
    return "\n".join(satirlar)


def _pace_hesapla(mesafe_km, sure_dk):
    if not mesafe_km or mesafe_km <= 0:
        return None
    pace_dk_km = sure_dk / mesafe_km
    dk = int(pace_dk_km)
    sn = int(round((pace_dk_km - dk) * 60))
    return f"{dk}:{sn:02d} dk/km"


def strava_kosu_pace_ozeti(access_token, kac_tane=8):
    """Son koşularının pace (tempo) verilerini özetler — antrenman
    önerilerinde 'gerçek pace'ine göre' konuşabilmek için kullanılır."""
    aktiviteler = strava_son_aktiviteleri_getir(access_token, kac_tane=30)
    kosular = [a for a in aktiviteler if a.get("type") in ("Run", "TrailRun", "VirtualRun")][:kac_tane]
    if not kosular:
        return ""

    satirlar = ["Son koşularımın pace (tempo) verileri:"]
    tum_pace_degerleri = []
    for a in kosular:
        mesafe_km = (a.get("distance", 0) or 0) / 1000
        sure_dk = (a.get("moving_time", 0) or 0) / 60
        pace = _pace_hesapla(mesafe_km, sure_dk)
        ort_nabiz = a.get("average_heartrate")
        tarih = (a.get("start_date_local", "") or "")[:10]
        if pace:
            satir = f"- {tarih}: {mesafe_km:.1f} km, pace {pace}"
            if ort_nabiz:
                satir += f", ort. nabız {ort_nabiz:.0f}"
            satirlar.append(satir)
            tum_pace_degerleri.append(sure_dk / mesafe_km if mesafe_km else None)

    gecerli_paceler = [p for p in tum_pace_degerleri if p]
    if gecerli_paceler:
        ort_pace = sum(gecerli_paceler) / len(gecerli_paceler)
        en_iyi_pace = min(gecerli_paceler)
        dk_ort, sn_ort = int(ort_pace), int(round((ort_pace - int(ort_pace)) * 60))
        dk_iyi, sn_iyi = int(en_iyi_pace), int(round((en_iyi_pace - int(en_iyi_pace)) * 60))
        satirlar.append(f"\nOrtalama pace: {dk_ort}:{sn_ort:02d} dk/km")
        satirlar.append(f"En iyi (en hızlı) pace: {dk_iyi}:{sn_iyi:02d} dk/km")

    return "\n".join(satirlar)


def intervals_kosu_pace_ozeti(api_key, kac_tane=8):
    """strava_kosu_pace_ozeti ile AYNI mantık — ama Intervals.icu için.
    Artık antrenmanla ilgili her şeyi (ağırlık + koşu + pace) TEK
    kaynaktan (Intervals.icu) çekebilmek için eklendi."""
    aktiviteler = intervals_aktiviteleri_getir(api_key, gun_sayisi=30)
    kosular = [a for a in aktiviteler if a.get("type") in ("Run", "TrailRun", "VirtualRun")][:kac_tane]
    if not kosular:
        return ""

    satirlar = ["Son koşularımın pace (tempo) verileri (Intervals.icu):"]
    tum_pace_degerleri = []
    for a in kosular:
        mesafe_km = (a.get("distance", 0) or 0) / 1000
        sure_dk = (a.get("moving_time", 0) or 0) / 60
        pace = _pace_hesapla(mesafe_km, sure_dk)
        ort_nabiz = a.get("icu_average_heart_rate") or a.get("average_heartrate")
        tarih = (a.get("start_date_local", "") or "")[:10]
        if pace:
            satir = f"- {tarih}: {mesafe_km:.1f} km, pace {pace}"
            if ort_nabiz:
                satir += f", ort. nabız {ort_nabiz:.0f}"
            satirlar.append(satir)
            tum_pace_degerleri.append(sure_dk / mesafe_km if mesafe_km else None)

    gecerli_paceler = [p for p in tum_pace_degerleri if p]
    if gecerli_paceler:
        ort_pace = sum(gecerli_paceler) / len(gecerli_paceler)
        en_iyi_pace = min(gecerli_paceler)
        dk_ort, sn_ort = int(ort_pace), int(round((ort_pace - int(ort_pace)) * 60))
        dk_iyi, sn_iyi = int(en_iyi_pace), int(round((en_iyi_pace - int(en_iyi_pace)) * 60))
        satirlar.append(f"\nOrtalama pace: {dk_ort}:{sn_ort:02d} dk/km")
        satirlar.append(f"En iyi (en hızlı) pace: {dk_iyi}:{sn_iyi:02d} dk/km")

    return "\n".join(satirlar)


_KOSU_ANAHTAR_KELIMELER = [
    "pace", "tempo", "interval", "koşu hız", "km/saat", "dakika/km",
    "dk/km", "hangi hızla", "ne hızla", "koşu antrenman",
    "koş",  # kök eşleşme: koşu, koşacağım, koşmalı, koşarken, koşuyorum, vb.
]


def _kosu_sorusu_mu(soru):
    soru_kucuk = soru.lower()
    return any(k in soru_kucuk for k in _KOSU_ANAHTAR_KELIMELER)


# ============== INTERVALS.ICU ENTEGRASYONU (UYKU/HRV) ==============
def intervals_baglantisini_kaydet(kullanici_id, api_key):
    if not DATABASE_URL:
        return
    baglanti = psycopg2.connect(DATABASE_URL)
    baglanti.autocommit = True
    imlec = baglanti.cursor()
    imlec.execute(
        "INSERT INTO intervals_baglantilar (kullanici_id, api_key) VALUES (%s, %s) "
        "ON CONFLICT (kullanici_id) DO UPDATE SET api_key = %s",
        (kullanici_id, api_key, api_key),
    )
    imlec.close()
    baglanti.close()


def intervals_api_key_getir(kullanici_id):
    if not DATABASE_URL:
        return None
    baglanti = psycopg2.connect(DATABASE_URL)
    imlec = baglanti.cursor()
    imlec.execute("SELECT api_key FROM intervals_baglantilar WHERE kullanici_id = %s", (kullanici_id,))
    satir = imlec.fetchone()
    imlec.close()
    baglanti.close()
    return satir[0] if satir else None


def intervals_uyku_verisi_getir(api_key, tarih=None):
    """Belirtilen tarihin (varsayılan bugün) uyku/HRV/toparlanma verisini
    çeker. athlete id '0' kullanılıyor — API anahtarına göre otomatik
    kendi hesabını bulur, ayrıca sporcu ID'si sormaya gerek kalmıyor."""
    if not tarih:
        tarih = _turkiye_simdi().strftime("%Y-%m-%d")
    yanit = requests.get(
        f"https://intervals.icu/api/v1/athlete/0/wellness/{tarih}",
        auth=("API_KEY", api_key),
        timeout=15,
    )
    yanit.raise_for_status()
    return yanit.json()


def intervals_ozet_metni(veri):
    if not veri:
        return ""
    uyku_sn = veri.get("sleepSecs") or veri.get("sleep_secs")
    uyku_skoru = veri.get("sleepScore") or veri.get("sleep_score")
    dinlenik_nabiz = veri.get("restingHR") or veri.get("resting_hr")
    hrv = veri.get("hrv")
    yorgunluk = veri.get("fatigue")
    agri = veri.get("soreness")
    stres = veri.get("stress")
    hazirlik = veri.get("readiness")

    satirlar = ["Bugünkü uyku/toparlanma verilerim (Huawei Band'den):"]
    if uyku_sn:
        saat = uyku_sn / 3600
        satirlar.append(f"- Uyku süresi: {saat:.1f} saat")
    if uyku_skoru:
        satirlar.append(f"- Uyku skoru: {uyku_skoru}")
    if dinlenik_nabiz:
        satirlar.append(f"- Dinlenik nabız: {dinlenik_nabiz}")
    if hrv:
        satirlar.append(f"- HRV: {hrv}")
    if yorgunluk:
        satirlar.append(f"- Yorgunluk: {yorgunluk}/5")
    if agri:
        satirlar.append(f"- Kas ağrısı: {agri}/5")
    if stres:
        satirlar.append(f"- Stres: {stres}")
    if hazirlik:
        satirlar.append(f"- Hazırlık (readiness): {hazirlik}")

    return "\n".join(satirlar) if len(satirlar) > 1 else ""


def intervals_aktiviteleri_getir(api_key, gun_sayisi=7):
    """Intervals.icu'daki tamamlanmış aktiviteleri (Huawei native
    bağlantısından gelenler dahil — ağırlık antrenmanı gibi Strava'nın
    yakalayamadığı türler burada olabilir) çeker."""
    bugun = _turkiye_simdi().date()
    baslangic = bugun - timedelta(days=gun_sayisi)
    yanit = requests.get(
        "https://intervals.icu/api/v1/athlete/0/activities",
        auth=("API_KEY", api_key),
        params={"oldest": baslangic.isoformat(), "newest": bugun.isoformat()},
        timeout=15,
    )
    yanit.raise_for_status()
    return yanit.json()


def intervals_aktiviteyi_metne_cevir(aktivite):
    isim = aktivite.get("name", "Aktivite")
    tur = aktivite.get("type", "")
    mesafe_km = (aktivite.get("distance", 0) or 0) / 1000
    sure_dk = (aktivite.get("moving_time", 0) or 0) / 60
    ort_nabiz = aktivite.get("icu_average_heart_rate") or aktivite.get("average_heartrate")
    tarih = aktivite.get("start_date_local", "")

    satirlar = [f"Aktivite: {isim} ({tur})", f"Tarih: {tarih}", f"Süre: {sure_dk:.0f} dakika"]
    if mesafe_km > 0:
        satirlar.append(f"Mesafe: {mesafe_km:.2f} km")
    if ort_nabiz:
        satirlar.append(f"Ortalama nabız: {ort_nabiz:.0f}")
    return "\n".join(satirlar)


def intervals_aktivite_daha_once_bildirildi_mi(kullanici_id, aktivite_id):
    if not DATABASE_URL:
        return False
    baglanti = psycopg2.connect(DATABASE_URL)
    imlec = baglanti.cursor()
    imlec.execute(
        "SELECT 1 FROM intervals_bildirilen_aktiviteler WHERE kullanici_id = %s AND aktivite_id = %s",
        (kullanici_id, str(aktivite_id)),
    )
    sonuc = imlec.fetchone() is not None
    imlec.close()
    baglanti.close()
    return sonuc


def intervals_aktiviteyi_bildirildi_isaretle(kullanici_id, aktivite_id):
    if not DATABASE_URL:
        return
    baglanti = psycopg2.connect(DATABASE_URL)
    baglanti.autocommit = True
    imlec = baglanti.cursor()
    imlec.execute(
        "INSERT INTO intervals_bildirilen_aktiviteler (kullanici_id, aktivite_id) "
        "VALUES (%s, %s) ON CONFLICT DO NOTHING",
        (kullanici_id, str(aktivite_id)),
    )
    imlec.close()
    baglanti.close()


# ============== KALICI KULLANICI PROFİLİ ==============
def profili_oku(kullanici_id):
    if not DATABASE_URL:
        return ""
    try:
        baglanti = psycopg2.connect(DATABASE_URL)
        imlec = baglanti.cursor()
        imlec.execute("SELECT profil FROM kullanici_profil WHERE kullanici_id = %s", (kullanici_id,))
        satir = imlec.fetchone()
        imlec.close()
        baglanti.close()
        return satir[0] if satir else ""
    except Exception:
        return ""


def tam_profili_olustur(kullanici_id):
    """profili_oku'nun TEK BAŞINA yetersiz kaldığı yerler için — temel
    profile ek olarak vücut ölçümü, beslenme ve ANTRENMAN GÜNLÜĞÜ
    özetlerini de birleştirip TEK bir metin döndürür. Bu, örneğin
    kullanıcı /antrenman_ekle ile bir hareketi zaten kaydettiyse,
    otomatik bildirim/yorum üreten fonksiyonların bunu GÖRMESİNİ sağlar
    — aksi halde zaten bilinen bir şeyi tekrar tekrar sormaya devam
    ederler (gerçek kullanımda tekrarlayan bir sorun olarak görüldü)."""
    profil = profili_oku(kullanici_id)
    olcum_ozeti = olcum_ozeti_ve_trend(kullanici_id)
    if olcum_ozeti:
        profil = (profil + "\n\n" + olcum_ozeti).strip() if profil else olcum_ozeti
    beslenme_ozeti = beslenme_gunluk_ozet_metni(kullanici_id)
    if beslenme_ozeti:
        profil = (profil + "\n\n" + beslenme_ozeti).strip() if profil else beslenme_ozeti
    antrenman_ozeti = antrenman_gunlugu_ozeti(kullanici_id)
    if antrenman_ozeti:
        profil = (profil + "\n\n" + antrenman_ozeti).strip() if profil else antrenman_ozeti
    return profil


def profili_yaz(kullanici_id, yeni_profil):
    if not DATABASE_URL:
        return
    baglanti = psycopg2.connect(DATABASE_URL)
    baglanti.autocommit = True
    imlec = baglanti.cursor()
    imlec.execute(
        "INSERT INTO kullanici_profil (kullanici_id, profil) VALUES (%s, %s) "
        "ON CONFLICT (kullanici_id) DO UPDATE SET profil = %s",
        (kullanici_id, yeni_profil, yeni_profil),
    )
    imlec.close()
    baglanti.close()


def profili_otomatik_guncelle(client_gemini, kullanici_id, soru, cevap):
    """Her konuşmadan sonra, kalıcı olarak hatırlanması gereken YENİ bir
    bilgi (hedef, kilo, sakatlık, tercih vb.) geçip geçmediğini kontrol
    eder, varsa profile ekler. Ucuz/hızlı bir model kullanır."""
    mevcut_profil = profili_oku(kullanici_id)
    talimat = (
        "Aşağıda bir kullanıcı ile antrenörü arasındaki SON mesaj çifti var. "
        "Kullanıcının MEVCUT PROFİLİ de altta. Eğer bu son mesajlarda, "
        "profile eklenmeye değer YENİ ve KALICI bir kişisel bilgi geçiyorsa "
        "(hedef, kilo, boy, yaş, sakatlık/kısıtlama, tercih, hedef yarış "
        "tarihi vb. — GEÇİCİ/günlük şeyler değil) profili güncelleyip TAM "
        "HALİNİ döndür. Yeni bir şey yoksa, SADECE 'DEĞİŞİKLİK_YOK' yaz. "
        "Profil kısa, madde madde, Türkçe olmalı, 15 satırı geçmemeli.\n\n"
        f"MEVCUT PROFİL:\n{mevcut_profil or '(henüz boş)'}\n\n"
        f"SON MESAJLAR:\nKullanıcı: {soru}\nAntrenör: {cevap}"
    )
    try:
        yanit = client_gemini.models.generate_content(
            model="gemini-flash-latest", contents=talimat,
        )
        sonuc = (yanit.text or "").strip()
        if sonuc and "DEĞİŞİKLİK_YOK" not in sonuc.upper():
            profili_yaz(kullanici_id, sonuc)
    except Exception as e:
        print(f"Profil güncellenirken hata: {e}")


# ============== ANTRENMAN GÜNLÜĞÜ (KISA, KALICI, GÜNLÜK KAYIT) ==============
def antrenman_kaydet(kullanici_id, aciklama):
    if not DATABASE_URL:
        return
    turkiye_tarihi = _turkiye_simdi().date()
    baglanti = psycopg2.connect(DATABASE_URL)
    baglanti.autocommit = True
    imlec = baglanti.cursor()
    imlec.execute(
        "INSERT INTO antrenman_gunlugu (kullanici_id, tarih, aciklama) VALUES (%s, %s, %s)",
        (kullanici_id, turkiye_tarihi, aciklama),
    )
    imlec.close()
    baglanti.close()


def antrenman_kaydet_tekrarsiz(kullanici_id, aciklama, benzersiz_anahtar):
    """antrenman_kaydet ile AYNI ama önce, verilen benzersiz_anahtar
    (genelde aktivitenin tam ISO zaman damgası, örn. '2026-08-17T14:16:29')
    bugünün kayıtlarında ZATEN var mı diye kontrol eder. Varsa tekrar
    yazmaz — aynı aktivitenin Strava/Intervals.icu/son_antrenman gibi
    farklı kaynaklardan tekrar tekrar kaydedilmesini önler."""
    if not DATABASE_URL:
        return
    bugun = _turkiye_simdi().date()
    baglanti = psycopg2.connect(DATABASE_URL)
    imlec = baglanti.cursor()
    imlec.execute(
        "SELECT 1 FROM antrenman_gunlugu WHERE kullanici_id = %s AND tarih = %s "
        "AND aciklama LIKE %s LIMIT 1",
        (kullanici_id, bugun, f"%{benzersiz_anahtar}%"),
    )
    zaten_var = imlec.fetchone() is not None
    imlec.close()
    baglanti.close()
    if not zaten_var:
        antrenman_kaydet(kullanici_id, aciklama)


def antrenman_gunlugu_ozeti(kullanici_id, gun_sayisi=4):
    """Son birkaç günün antrenman özetini kısa metin olarak döndürür —
    konuşma geçmişi ne kadar dolu olursa olsun, 'dün ne yaptık' sorusuna
    HER ZAMAN doğru cevap verilebilmesi için."""
    if not DATABASE_URL:
        return ""
    bugun = _turkiye_simdi().date()
    baslangic = bugun - timedelta(days=gun_sayisi)
    baglanti = psycopg2.connect(DATABASE_URL)
    imlec = baglanti.cursor()
    imlec.execute(
        "SELECT tarih, aciklama FROM antrenman_gunlugu WHERE kullanici_id = %s "
        "AND tarih >= %s ORDER BY tarih DESC, id DESC",
        (kullanici_id, baslangic),
    )
    satirlar = imlec.fetchall()
    imlec.close()
    baglanti.close()
    if not satirlar:
        return ""

    metin_satirlari = ["Son günlerin GERÇEK antrenman günlüğüm (bu kayıtlara güven, tahmin etme):"]
    for tarih, aciklama in satirlar:
        fark = (bugun - tarih).days
        if fark == 0:
            etiket = "BUGÜN"
        elif fark == 1:
            etiket = "DÜN"
        else:
            etiket = f"{fark} gün önce"
        metin_satirlari.append(f"- {etiket} ({tarih.strftime('%d.%m')}): {aciklama}")
    return "\n".join(metin_satirlari)


def antrenman_gunlugunu_otomatik_guncelle(client_gemini, kullanici_id, soru, cevap):
    """profili_otomatik_guncelle'e paralel çalışır — bu sefer 'kalıcı
    profil bilgisi' değil, 'bugün/dün YAPILAN spesifik bir antrenman'
    var mı diye bakar, varsa kısa bir günlük satırı olarak kaydeder.
    Sadece son tek mesaja değil, SON BİRKAÇ mesaja da bakıyor — çünkü
    hareket/set/tekrar detayları genelde önceki mesajlarda konuşulup,
    'tamamdır bitti' gibi son mesaj tek başına yeterli bağlam içermez."""
    onceki_baglam = ""
    try:
        son_mesajlar = gecmisi_oku(kullanici_id, limit=8)
        satirlar = []
        for m in son_mesajlar:
            rol = "Kullanıcı" if m.get("role") == "user" else "Antrenör"
            metin = ""
            for p in m.get("parts", []):
                if "text" in p:
                    metin += p["text"]
            if metin:
                satirlar.append(f"{rol}: {metin[:300]}")
        onceki_baglam = "\n".join(satirlar[-8:])
    except Exception:
        pass

    bugunku_kayitlar = antrenman_gunlugu_ozeti(kullanici_id, gun_sayisi=1)

    talimat = (
        "Aşağıda bir kullanıcı ile antrenörü arasındaki SON KONUŞMA "
        "(önceki birkaç mesaj + en son mesaj çifti) var. Kullanıcı "
        "GERÇEKTEN YAPTIĞI/TAMAMLADIĞI bir antrenmandan (koşu, kuvvet "
        "antrenmanı vb.) bahsediyorsa — detaylar (hangi hareket, kaç "
        "set/tekrar) ÖNCEKİ mesajlarda geçmiş olsa bile — bunu TEK "
        "SATIRLIK kısa bir günlük kaydı olarak özetle, ÖNCEKİ mesajlardaki "
        "detayları da (hareket isimleri gibi) dahil et (örn. 'Bacak günü "
        "yaptı: Squat, RDL, BSS' ya da '6. hafta 1. antrenmanını (piramit "
        "koşu) tamamladı'). Sadece GERÇEKTEN YAPILDIĞI/BİTTİĞİ belirtilen "
        "antrenmanları kaydet, PLANLANAN/önerilen ama henüz yapılmamış "
        "antrenmanları YAZMA.\n\n"
        "🚨 ÖNEMLİ — TEKRAR KAYDETME: Aşağıda 'BUGÜN ZATEN KAYITLI OLANLAR' "
        "listesi var. Eğer bu konuşmadaki antrenman, o listede ZATEN "
        "(aynı ya da çok benzer şekilde) varsa, TEKRAR KAYDETME — SADECE "
        "'YOK' yaz. Sadece listede HİÇ olmayan, gerçekten YENİ bir bilgi "
        "varsa kaydet.\n\n"
        "Eğer kaydedilecek yeni bir şey yoksa SADECE 'YOK' yaz.\n\n"
        f"BUGÜN ZATEN KAYITLI OLANLAR:\n{bugunku_kayitlar or '(henüz hiçbir şey kayıtlı değil)'}\n\n"
        f"ÖNCEKİ MESAJLAR (bağlam için):\n{onceki_baglam}\n\n"
        f"EN SON MESAJ ÇİFTİ:\nKullanıcı: {soru}\nAntrenör: {cevap}"
    )
    try:
        yanit = client_gemini.models.generate_content(
            model="gemini-flash-latest", contents=talimat,
        )
        sonuc = (yanit.text or "").strip()
        if sonuc and "YOK" != sonuc.upper() and len(sonuc) > 5:
            antrenman_kaydet(kullanici_id, sonuc)
    except Exception as e:
        print(f"Antrenman günlüğü güncellenirken hata: {e}")


# ============== TARİH/SAAT ==============
def _turkiye_simdi():
    try:
        from zoneinfo import ZoneInfo
        return datetime.now(ZoneInfo("Europe/Istanbul"))
    except Exception:
        return datetime.now()


def bugunun_tarihi():
    simdi = _turkiye_simdi()
    return f"{simdi.day} {AYLAR_TR[simdi.month - 1]} {simdi.year}, {GUNLER_TR[simdi.weekday()]}"


def su_anki_saat():
    simdi = _turkiye_simdi()
    saat = simdi.hour
    if 5 <= saat < 12:
        vakit = "sabah"
    elif 12 <= saat < 17:
        vakit = "öğleden sonra"
    elif 17 <= saat < 21:
        vakit = "akşam"
    else:
        vakit = "gece"
    return simdi.strftime("%H:%M"), vakit


# ============== SES YAZIYA ÇEVİRME ==============
def ses_yaziya_cevir(client_gemini, ses_bytes, mime_tipi="audio/ogg"):
    ses_b64 = base64.b64encode(ses_bytes).decode("utf-8")
    denenecek_tipler = [mime_tipi] + [
        t for t in ["audio/ogg", "audio/wav", "audio/mp3", "audio/mp4", "audio/webm"]
        if t != mime_tipi
    ]
    for tip in denenecek_tipler:
        for model_adi in GEMINI_MODEL_LISTESI:
            try:
                yanit = client_gemini.models.generate_content(
                    model=model_adi,
                    contents=[
                        {"role": "user", "parts": [
                            {"text": "Bu ses kaydını sadece yazıya çevir, başka hiçbir şey ekleme."},
                            {"inline_data": {"mime_type": tip, "data": ses_b64}},
                        ]}
                    ],
                )
                if yanit.text and yanit.text.strip():
                    return yanit.text.strip()
            except Exception:
                continue
    return None


_TTS_MODEL_LISTESI = ["gemini-2.5-flash-preview-tts", "gemini-2.5-pro-preview-tts"]


def metni_sese_cevir(client_gemini, metin, ses_adi="Kore"):
    """Metni Gemini'nin sesli üretim (TTS) modeliyle sese çevirir,
    WAV formatında ses byte'ları döndürür. Uzun metinlerde ilk ~600
    karakterle sınırlar (TTS modelleri uzun metinlerde bozulabiliyor)."""
    import wave

    kisa_metin = metin[:600] if len(metin) > 600 else metin

    for model_adi in _TTS_MODEL_LISTESI:
        try:
            yanit = client_gemini.models.generate_content(
                model=model_adi,
                contents=kisa_metin,
                config={
                    "response_modalities": ["AUDIO"],
                    "speech_config": {
                        "voice_config": {
                            "prebuilt_voice_config": {"voice_name": ses_adi}
                        }
                    },
                },
            )
            pcm_veri = yanit.candidates[0].content.parts[0].inline_data.data
            tampon = BytesIO()
            with wave.open(tampon, 'wb') as wf:
                wf.setnchannels(1)
                wf.setsampwidth(2)
                wf.setframerate(24000)
                wf.writeframes(pcm_veri)
            return tampon.getvalue()
        except Exception as e:
            print(f"TTS hatası ({model_adi}): {e}")
            continue
    return None


# ============== TELEGRAM MESAJ UZUNLUK GÜVENLİĞİ ==============
TELEGRAM_MAX_UZUNLUK = 3900  # Telegram'ın ~4096 sınırına güvenli pay


async def guvenli_reply(message, metin, **kwargs):
    """update.message.reply_text yerine kullanılır — mesaj çok uzunsa
    (Telegram'ın 4096 karakter sınırını aşarsa) otomatik olarak
    parçalara bölüp sırayla gönderir, hata fırlatmaz."""
    if not metin:
        metin = "(boş cevap)"
    if len(metin) <= TELEGRAM_MAX_UZUNLUK:
        await message.reply_text(metin, **kwargs)
        return
    parcalar = [metin[i:i + TELEGRAM_MAX_UZUNLUK] for i in range(0, len(metin), TELEGRAM_MAX_UZUNLUK)]
    for i, parca in enumerate(parcalar):
        if i == len(parcalar) - 1:
            await message.reply_text(parca, **kwargs)  # butonlar/ekstra son parçada
        else:
            await message.reply_text(parca)


async def guvenli_send_message(bot, chat_id, metin, **kwargs):
    """context.bot.send_message için aynı güvenlik, proaktif mesajlarda kullanılır."""
    if not metin:
        metin = "(boş cevap)"
    if len(metin) <= TELEGRAM_MAX_UZUNLUK:
        await bot.send_message(chat_id=chat_id, text=metin, **kwargs)
        return
    parcalar = [metin[i:i + TELEGRAM_MAX_UZUNLUK] for i in range(0, len(metin), TELEGRAM_MAX_UZUNLUK)]
    for parca in parcalar:
        await bot.send_message(chat_id=chat_id, text=parca)



def cevap_uret(client_gemini, soru, baglam, gecmis, gorsel_b64=None, gorsel_mime=None,
                yumusak=False, profil=""):
    bugun = bugunun_tarihi()
    saat, vakit = su_anki_saat()

    ton_talimati = (
        "TON: Benimle samimi, sıcak, motive edici ve gerçek bir antrenör "
        "gibi konuş — resmi/mesafeli bir asistan gibi değil."
        if not yumusak else
        "TON: Yumuşak, nazik, destekleyici ve sabırlı bir dille konuş — "
        "baskıcı/sert bir koç gibi değil, anlayışlı bir rehber gibi."
    )

    format_kurali = (
        "📱 FORMAT — TELEGRAM MESAJI YAZ, RAPOR YAZMA: Sen bir arkadaşının/"
        "antrenörünün Telegram'dan sana attığı bir MESAJ yazıyorsun, resmi "
        "bir rapor/analiz belgesi DEĞİL. Şu kalıplardan KAÇIN:\n"
        "- '### Başlık' gibi markdown başlıkları KULLANMA — gerçek biri "
        "sana mesaj atarken başlık açmaz.\n"
        "- Her cevabı madde madde (bullet point) listeler halinde bölmek "
        "yerine, çoğu zaman AKICI, doğal cümlelerle (normal konuşma gibi) "
        "yaz — madde listesi sadece gerçekten sıralı bir şey (egzersiz "
        "listesi gibi) varsa kullanılsın.\n"
        "- **Kalın yazıyı** her cümlede değil, sadece GERÇEKTEN vurgulanması "
        "gereken 1-2 kelimede kullan.\n"
        "- HİTAP ÇEŞİTLİLİĞİ: 'Şampiyon' ya da 'Dostum' gibi bir hitabı HER "
        "mesajda kullanma — çoğu mesajda HİÇ hitap kullanma (direkt konuya "
        "gir), ara sıra kullanacaksan da çeşitlendir. Aynı kelimeyi sürekli "
        "tekrarlamak, samimi değil ROBOTİK/ŞABLON hissi veriyor.\n"
        "- EMOJİ: Mesaj başına EN FAZLA 1-2 emoji kullan (hiç kullanmamak "
        "da tamamen normal). ASLA aynı emoji kombinasyonunu (örn. "
        "'🚀🏃‍♂️💥' gibi) tekrar tekrar, birebir aynı şekilde mesaj sonuna "
        "ekleme — bu, en net 'bu bir şablon/bot' işaretlerinden biri. "
        "Gerçek bir insan her mesajının sonuna aynı 3 emojiyi koymaz.\n"
        "- Kısa bir soruya kısa cevap ver — her mesajı bir 'analiz raporuna' "
        "çevirme, sadece konu gerçekten karmaşıksa uzun/yapılandırılmış yaz."
    )

    profil_blogu = (
        f"📋 KULLANICI HAKKINDA KALICI BİLGİLER (bunlar her zaman doğrudur, "
        f"konuşma ne kadar eski olursa olsun unutma):\n{profil}\n\n"
        if profil else ""
    )

    sistem_mesaji = (
        "🤫 GENEL KURAL — İÇ MUHAKEMENİ SESSİZ TUT: Arşivden gelen notlar "
        "arasında konuyla (yemek, antrenman, program, ne olursa olsun) "
        "ALAKASIZ olanlar varsa, bunları SESSİZCE görmezden gel ve "
        "cevabını sadece alakalı bilgiye dayandır. ASLA 'video notları "
        "alakasız, onları yok sayıyorum', 'bu bilgi başkalarına ait, "
        "kullanmıyorum' gibi bir açıklamayla cevaba BAŞLAMA ya da bunu "
        "belirtme — bu, kullanıcı için gereksiz ve tekrarlayıcı geliyor. "
        "Notların hangilerinin alakalı/alakasız olduğunu KENDİ İÇİNDE "
        "değerlendir, sonucu sessizce uygula, sadece nihai/alakalı "
        "cevabı ver.\n\n"
        "🏷️ ÖNCEKİ MESAJLARDAKİ '[...]' ETİKETLERİNİ DOĞRU OKU: Geçmiş "
        "mesajlarda '[FOTOĞRAF ANALİZİ — BEN ürettim...]' gibi köşeli "
        "parantezli bir not görürsen, bu SENİN (antrenörün) kendi "
        "tahminin/analizin demektir — kullanıcı bu sayıları SANA "
        "SÖYLEMEDİ, sen fotoğraf/veriye bakarak ÜRETTİN. Kullanıcı "
        "sonradan 'bu rakamlar nereden geldi' diye sorarsa, 'sen "
        "söylemiştin' deme — dürüstçe 'ben fotoğrafa bakarak tahmin "
        "etmiştim, kesin doğru olmayabilir' de.\n\n"
        "🚨 TEMEL KURAL — GERÇEK VERİYE DAYAN, UYDURMA: Aşağıdaki üç "
        "durumda ASLA tahmin/uydurma yapma, sadece verilen gerçek bilgiyi "
        "kullan, yoksa dürüstçe 'elimde net yok' de:\n"
        "1) KİŞİSEL GEÇMİŞ: 'geçen hafta ne yaptık', 'hatırlıyor musun', "
        "'dün ne yaptık' gibi sorularda SADECE ÖNCEKİ MESAJLAR'da, "
        "'📋 KALICI BİLGİLER'de, 'Son günlerin GERÇEK antrenman günlüğüm...' "
        "notunda ya da '[GERÇEK KİŞİSEL GEÇMİŞ...]' etiketli notlarda "
        "GERÇEKTEN yazanı kullan — antrenman günlüğü notu varsa buna "
        "TAM GÜVEN, bu gerçek ve doğrulanmış bir kayıt. '[Genel video "
        "içeriği...]' notları BAŞKA insanların hikayeleri, asla "
        "kullanıcınınmış gibi anlatma.\n"
        "2) SPESİFİK PROGRAM/İÇERİK: Bir programın (örn. '8 haftada 5K') "
        "tam hafta/gün detayını sorduğunda, SADECE notlarda gerçekten "
        "yazıyorsa kesin bilgi ver. Yoksa 'genel mantık şöyle ama tam "
        "detay elimde yok' diye belirsizliği açıkça söyle.\n"
        "3) PACE/TEMPO: '[GERÇEK STRAVA VERİSİ...]' verilmişse öneriyi "
        "buna dayandır. Kilo/boy gibi genel özelliklerden soyut pace "
        "tahmini ('100 kg birisin, böyle koşarsın' gibi) ASLA uydurma — "
        "veri yoksa sor.\n"
        "4) UYKU/TOPARLANMA: 'Bugünkü uyku/toparlanma verilerim' notu "
        "varsa, antrenman önerisini buna göre uyarla (örn. uyku kötüyse "
        "hafiflet/dinlenme öner). Bu veri yoksa uyku hakkında tahmin "
        "yürütme, sadece genel tavsiye ver. Sensör verileri (uyku, "
        "nabız, HRV) GÜN İÇİNDE GÜNCELLENEBİLİR (örn. erken uyanıp "
        "tekrar uyuma) — eğer şu an çekilen taze veri, bugün daha önce "
        "konuşulan bir rakamdan farklıysa bu 'tutarsızlık' DEĞİL, "
        "sadece güncelleme demektir; EN TAZE veriye güven, farkı "
        "sorgulayıp kafa karıştırma.\n\n"
        "🚨🚨 ISRAR ALTINDA DA UYDURMA YAPMA — EN KRİTİK KURAL: Kullanıcı "
        "'bunu zaten bilmen lazım', 'daha önce konuşmuştuk', 'bunu neden "
        "bilmiyorsun' diye ISRAR EDİP SENİ SIKIŞTIRSA BİLE, eğer o bilgi "
        "GERÇEKTEN ÖNCEKİ MESAJLAR'da ya da notlarda YOKSA, ASLA 'evet "
        "hatırlıyorum, bu gerçek bir bilgi, uydurmuyorum' diye SAHTE BİR "
        "GÜVEN İÇİNDE cevap uydurma. Kullanıcıyı memnun etmek için "
        "dürüstlüğünden ASLA vazgeçme — ısrar ne kadar güçlü olursa olsun, "
        "gerçekten elinde olmayan bir bilgiyi 'elimde var' gibi sunmak, "
        "'elimde yok' demekten çok daha kötü bir hatadır. Bu durumda sakin "
        "ve net kal: 'Israrını anlıyorum ama gerçekten önceki kayıtlarımda "
        "bunu bulamıyorum, bana tekrar gönderir/hatırlatır mısın?' de. "
        "Kullanıcı sinirlense de üzülse de, UYDURMAK yerine DÜRÜST KALMAK "
        "her zaman doğru seçimdir.\n\n"
        "📅 TARİH ALGISI: ÖNCEKİ MESAJLAR'daki [BUGÜN]/[DÜN]/[X gün önce] "
        "etiketlerine uy — DÜN'kü bir olayı ('bugün doğum günüm' gibi) şu "
        "anki bugünmüş gibi ele alma.\n\n"
        "📌 KAYNAK ÇELİŞKİSİ: Farklı video_id'ler çelişiyorsa görmezden "
        "gelme, 'bazı kaynaklar şöyle, bazıları böyle' diye belirt.\n\n"
        "📆 'X. GÜN' BELİRSİZLİĞİ: Videolar günleri bazen sayıyla (1. gün, "
        "2. gün), bazen isimle (Pazartesi) anlatıyor, tutarsız olabilir. "
        "Notlarda net tanımlı değilse tahmin etme, kullanıcıya hangi günü "
        "kastettiğini sor. AYRICA çok önemli bir ayrım: bir video "
        "'1. antrenman/1. seans, 2. antrenman/2. seans' diye SIRALI "
        "ANTRENMAN GÜNLERİNDEN bahsediyorsa (dinlenme günleri hariç, "
        "sadece koşu/kuvvet günleri sayılmış), bunu ASLA art arda gelen "
        "takvim günleriymiş gibi ('1. gün, 2. gün, 3. gün' diye peş peşe) "
        "sunma — aralarında dinlenme günleri olabilir. Bu durumda ya "
        "gerçek gün ismini (Pazartesi/Çarşamba/Cuma gibi) kullan ya da "
        "açıkça 'X. antrenman günü (aralarında dinlenme var)' diye belirt, "
        "asla 'X. gün' diye takvim günü izlenimi verme.\n\n"
        f"{profil_blogu}"
        f"Bugünün tarihi: {bugun}. Şu anki saat: {saat} ({vakit}). Bu bilgiyi "
        f"antrenman/beslenme önerilerinde dikkate al.\n\n"
        "Sen benim kişisel hybrid antrenörümsün. Amacın, beni hybrid "
        "sistemde (koşu, bisiklet, yüzme, kuvvet vb.) en iyi versiyonuma "
        "ulaştırmak için elinden gelen tüm yardımı sağlamak. Spor bilimi, "
        "antrenman periyotlaması, beslenme, toparlanma konularında "
        "deneyimli bir antrenör gibisin.\n\n"
        "Sana verilen notlar (varsa) senin kendi bilgi birikimin gibi "
        "davran, dışarıdan kaynak gibi sunma. Alakalıysa öncelikle bunlara "
        "dayan, alakasızsa kendi genel uzmanlığından cevap ver.\n\n"
        f"{ton_talimati}\n\n"
        f"{format_kurali}"
    )

    contents = list(gecmis)
    kullanici_mesaji = f"(Alakalı notlar varsa aşağıda, yoksa yok say):\n\n{baglam}\n\nSORU: {soru}"
    parts = [{"text": kullanici_mesaji}]
    if gorsel_b64:
        parts.append({"inline_data": {"mime_type": gorsel_mime, "data": gorsel_b64}})
    contents.append({"role": "user", "parts": parts})

    # 1) ÖNCE Claude'u dene (Opus -> Sonnet). Anthropic anahtarı yoksa
    # ya da ikisi de başarısız olursa, sessizce Gemini zincirine düşülür.
    if ANTHROPIC_API_KEY:
        claude_mesajlari = []
        for m in gecmis:
            rol = "assistant" if m.get("role") == "model" else "user"
            metin = ""
            for p in m.get("parts", []):
                if "text" in p:
                    metin += p["text"]
            if metin:
                claude_mesajlari.append({"role": rol, "content": metin})

        claude_icerik = [{"type": "text", "text": kullanici_mesaji}]
        if gorsel_b64:
            claude_icerik.append({
                "type": "image",
                "source": {"type": "base64", "media_type": gorsel_mime or "image/jpeg", "data": gorsel_b64},
            })
        claude_mesajlari.append({"role": "user", "content": claude_icerik})

        for model_adi in CLAUDE_MODEL_LISTESI:
            try:
                claude_client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY, timeout=45.0)
                yanit = claude_client.messages.create(
                    model=model_adi,
                    max_tokens=4096,
                    system=sistem_mesaji,
                    messages=claude_mesajlari,
                )
                metin = "".join(b.text for b in yanit.content if hasattr(b, "text"))
                if metin.strip():
                    print(f"✅ CEVAP ÜRETİLDİ: model='{model_adi}' (Claude)")
                    return metin
            except Exception as e:
                print(f"❌ '{model_adi}' (Claude) başarısız oldu, sıradakini deniyorum. Hata: {e}")
                continue
        print("⚠️ Tüm Claude denemeleri başarısız, Gemini zincirine düşülüyor.")

    # 2) Claude devre dışıysa ya da tamamen başarısız olduysa Gemini'ye düş
    for model_adi in ANA_CEVAP_MODEL_LISTESI:
        try:
            ayarlar = {"system_instruction": sistem_mesaji}
            if model_adi == "gemini-3.1-pro-preview":
                # Pro yoğunsa (503/deadline) uzun uzun beklemeyelim,
                # 10 saniye içinde cevap gelmezse hızlıca Flash'e düşelim.
                ayarlar["http_options"] = {"timeout": 10000}  # milisaniye
            yanit = client_gemini.models.generate_content(
                model=model_adi, contents=contents,
                config=ayarlar,
            )
            if yanit.text and yanit.text.strip():
                print(f"✅ CEVAP ÜRETİLDİ: model='{model_adi}'")
                return yanit.text
        except Exception as e:
            print(f"❌ '{model_adi}' başarısız oldu, sıradakini deniyorum. Hata: {e}")
            continue
    return "Şu an cevap üretemedim, lütfen tekrar dener misin?"


# ============== PROGRAM -> TAKVİM/EXCEL (kısaltılmış, arayuz.py'dekiyle aynı mantık) ==============
def programdan_json_cikar(client_gemini, program_metni):
    talimat = (
        "Aşağıdaki metinde bir antrenman/beslenme programı var. SADECE "
        "JSON döndür, başka hiçbir şey yazma:\n\n"
        '[{"gun": "Pazartesi", "baslik": "...", "baslangic_saat": "18:00", '
        '"bitis_saat": "19:00", "aciklama": "..."}]\n\n'
        "gun değeri: Pazartesi, Salı, Çarşamba, Perşembe, Cuma, Cumartesi, Pazar. "
        "Program yoksa boş liste [] döndür.\n\n"
        f"METİN:\n{program_metni}"
    )
    for model_adi in GEMINI_MODEL_LISTESI:
        try:
            yanit = client_gemini.models.generate_content(model=model_adi, contents=talimat)
            metin = re.sub(r"^```json\s*|\s*```$", "", yanit.text.strip(), flags=re.MULTILINE).strip("`").strip()
            return json.loads(metin)
        except Exception:
            continue
    return None


def sonraki_gun_tarihi(hedef_gun_index):
    bugun = datetime.now().date()
    fark = (hedef_gun_index - bugun.weekday()) % 7
    return bugun + timedelta(days=fark)


def ics_plan_olustur(etkinlikler):
    satirlar = ["BEGIN:VCALENDAR", "VERSION:2.0", "PRODID:-//Kocum//TR//", "CALSCALE:GREGORIAN"]
    for e in etkinlikler:
        gun = e.get("gun", "")
        if gun not in GUN_INDEX:
            continue
        try:
            bs_saat, bs_dk = map(int, e.get("baslangic_saat", "18:00").split(":"))
            bt_saat, bt_dk = map(int, e.get("bitis_saat", "19:00").split(":"))
        except Exception:
            continue
        tarih = sonraki_gun_tarihi(GUN_INDEX[gun])
        satirlar.append("BEGIN:VEVENT")
        satirlar.append(f"UID:{uuid.uuid4()}")
        satirlar.append(f"DTSTAMP:{datetime.now().strftime('%Y%m%dT%H%M%SZ')}")
        satirlar.append(f"DTSTART;TZID=Europe/Istanbul:{tarih.strftime('%Y%m%d')}T{bs_saat:02d}{bs_dk:02d}00")
        satirlar.append(f"DTEND;TZID=Europe/Istanbul:{tarih.strftime('%Y%m%d')}T{bt_saat:02d}{bt_dk:02d}00")
        satirlar.append(f"SUMMARY:[{UYGULAMA_ADI}] {e.get('baslik', 'Etkinlik')}")
        if e.get("aciklama"):
            satirlar.append(f"DESCRIPTION:{e['aciklama']}")
        satirlar.append("END:VEVENT")
    satirlar.append("END:VCALENDAR")
    return "\r\n".join(satirlar)


def programdan_excel_json_cikar(client_gemini, program_metni):
    talimat = (
        "Aşağıdaki metinde bir antrenman/beslenme programı var. SADECE "
        "JSON döndür:\n\n"
        '[{"gun": "Pazartesi", "kategori": "Antrenman", "baslik": "...", '
        '"detay": "...", "sure_kalori": "60 dk"}]\n\n'
        f"METİN:\n{program_metni}"
    )
    for model_adi in GEMINI_MODEL_LISTESI:
        try:
            yanit = client_gemini.models.generate_content(model=model_adi, contents=talimat)
            metin = re.sub(r"^```json\s*|\s*```$", "", yanit.text.strip(), flags=re.MULTILINE).strip("`").strip()
            return json.loads(metin)
        except Exception:
            continue
    return None


def excel_plan_olustur(satirlar):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Program"
    basliklar = ["Gün", "Kategori", "Başlık", "Detay", "Süre/Kalori"]
    b_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    b_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    h_font = Font(name="Arial", size=10)
    kenar = Border(*(Side(style="thin", color="D1D5DB"),) * 4)

    for i, m in enumerate(basliklar, start=1):
        h = ws.cell(row=1, column=i, value=m)
        h.font, h.fill, h.border = b_font, b_fill, kenar
        h.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for r, e in enumerate(satirlar, start=2):
        degerler = [e.get("gun", ""), e.get("kategori", ""), e.get("baslik", ""),
                    e.get("detay", ""), e.get("sure_kalori", "")]
        for c, d in enumerate(degerler, start=1):
            h = ws.cell(row=r, column=c, value=d)
            h.font, h.border = h_font, kenar
            h.alignment = Alignment(vertical="top", wrap_text=True)

    for i, g in enumerate([12, 12, 22, 45, 14], start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = g
    ws.freeze_panes = "A2"

    tampon = BytesIO()
    wb.save(tampon)
    return tampon.getvalue()


# ============== TELEGRAM HANDLER'LARI ==============
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        f"Merhaba! Ben {UYGULAMA_ADI}, senin kişisel hybrid antrenörünüm. 💪\n\n"
        f"Yazarak, sesli mesajla ya da fotoğraf göndererek soru sorabilirsin. "
        f"Eski sohbet (.json) ya da video transkripti (.md) dosyalarını da "
        f"gönderip arşive ekleyebilirsin.\n\n"
        f"📋 Profil:\n"
        f"/profil_olustur — birkaç soruyla tam profil oluştur (ÖNERİLİR)\n"
        f"/profil_goster — hakkında bildiklerim\n"
        f"/profil_ekle <bilgi> — kalıcı bir bilgi ekle\n"
        f"/profil_sil — profili sıfırla\n\n"
        f"📏 Vücut ölçümleri:\n"
        f"/olcum_ekle — ölçümlerini kaydet (kol, bel, kalça, göğüs, bacak, kilo)\n"
        f"/olcum_gecmisi — geçmiş ölçümlerini gör\n"
        f"/olcum_hatirlatma_ac, /olcum_hatirlatma_kapat — aylık hatırlatma\n\n"
        f"🍽️ Beslenme:\n"
        f"/yemek_ekle — yemek modunu aç/kapat (açıkken TÜM fotoğraflar otomatik kaydedilir)\n"
        f"/beslenme_ozet [gün] — beslenme özeti (varsayılan bugün)\n"
        f"/yemek_duzelt <kalori> <p> <k> <y> — son kaydı gerçek değerlerle düzelt\n\n"
        f"🔊 Sesli cevap:\n"
        f"/sesli_cevap_ac — cevapları sesli de al\n"
        f"/sesli_cevap_kapat — kapat\n\n"
        f"😴 Uyku/Toparlanma (Huawei Band):\n"
        f"/intervals_baglan <api_key> — Intervals.icu hesabını bağla\n"
        f"/uyku_durumu — bugünkü uyku/HRV verini yorumlat\n"
        f"/antrenman_gecmisi [gün] — son antrenmanlarının günlüğü (varsayılan 14 gün)\n"
        f"/antrenman_ekle <açıklama> — garantili manuel kayıt (otomatik kaçırırsa)\n\n"
        f"🏃 Strava:\n"
        f"/strava_baglan <token> — hesabını bağla\n"
        f"/son_antrenman — son aktiviteni yorumlat\n"
        f"/strava_ozet [gün] — trend özeti (varsayılan 7 gün)\n\n"
        f"☀️ Sabah mesajı:\n"
        f"/sabah_ac — her sabah 07:00'de otomatik mesaj\n"
        f"/sabah_kapat — kapat\n\n"
        f"🔍 Arşiv araçları:\n"
        f"/video_var_mi <id> — bir video arşivde var mı kontrol et\n"
        f"/zorla_video <id> — bir videoyu garanti kullan\n\n"
        f"⚙️ Genel:\n"
        f"/id — Telegram ID'ni gösterir\n"
        f"/yumusak_ac, /yumusak_kapat — ton ayarı\n"
        f"/temizle — sohbet geçmişini sıfırla\n"
        f"/web_sohbetlerini_getir <email> — eski web sohbetlerini taşı"
    )


async def id_goster(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(f"Telegram ID'n: {update.effective_user.id}")


async def model_listesi(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Geçici tanı komutu — senin API anahtarına GERÇEKTEN hangi
    modellerin açık olduğunu, tahmin etmeden doğrudan Google'dan sorar."""
    client_gemini, _ = istemcileri_al()
    try:
        modeller = client_gemini.models.list()
        satirlar = []
        for m in modeller:
            ad = getattr(m, "name", str(m))
            yontemler = getattr(m, "supported_actions", None) or getattr(m, "supported_generation_methods", None) or []
            if "generateContent" in yontemler or "generate_content" in [str(y).lower() for y in yontemler] or not yontemler:
                if "pro" in ad.lower() or "flash" in ad.lower():
                    satirlar.append(ad)
        if not satirlar:
            satirlar = [getattr(m, "name", str(m)) for m in client_gemini.models.list()]
        await guvenli_reply(update.message, "📋 Sana açık modeller:\n" + "\n".join(satirlar))
    except Exception as e:
        await update.message.reply_text(f"Model listesi alınamadı: {e}")


async def yumusak_ac(update: Update, context: ContextTypes.DEFAULT_TYPE):
    yumusak_ton_ayarla(update.effective_user.id, True)
    await update.message.reply_text("Tamamdır, bundan sonra daha yumuşak bir tonla konuşacağım. 🌿")


async def yumusak_kapat(update: Update, context: ContextTypes.DEFAULT_TYPE):
    yumusak_ton_ayarla(update.effective_user.id, False)
    await update.message.reply_text("Tamamdır, normal (motive edici) tona döndüm. 💪")


async def temizle(update: Update, context: ContextTypes.DEFAULT_TYPE):
    kullanici_id = update.effective_user.id
    if DATABASE_URL:
        try:
            baglanti = psycopg2.connect(DATABASE_URL)
            baglanti.autocommit = True
            imlec = baglanti.cursor()
            imlec.execute("DELETE FROM tg_mesajlar WHERE kullanici_id = %s", (kullanici_id,))
            imlec.close()
            baglanti.close()
        except Exception:
            pass
    await update.message.reply_text("Sohbet geçmişin temizlendi, sıfırdan başlıyoruz.")


async def profil_goster(update: Update, context: ContextTypes.DEFAULT_TYPE):
    kullanici_id = update.effective_user.id
    profil = profili_oku(kullanici_id)
    if not profil:
        await update.message.reply_text(
            "Henüz kalıcı bir profil bilgin yok. Sohbet ettikçe otomatik "
            "oluşacak, ya da /profil_ekle ile elle ekleyebilirsin."
        )
        return
    await update.message.reply_text(f"📋 Senin hakkında bildiklerim:\n\n{profil}")


async def profil_ekle(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not context.args:
        await update.message.reply_text(
            "Kullanım: /profil_ekle <bilgi>\nÖrn: /profil_ekle Sol dizimde eski bir sakatlık var, ağır squat yapamıyorum"
        )
        return
    kullanici_id = update.effective_user.id
    yeni_bilgi = " ".join(context.args)
    mevcut = profili_oku(kullanici_id)
    guncel = (mevcut + "\n- " + yeni_bilgi).strip() if mevcut else "- " + yeni_bilgi
    profili_yaz(kullanici_id, guncel)
    await update.message.reply_text("✅ Profiline eklendi, bundan sonra bunu hep hatırlayacağım.")


async def profil_sil(update: Update, context: ContextTypes.DEFAULT_TYPE):
    kullanici_id = update.effective_user.id
    profili_yaz(kullanici_id, "")
    await update.message.reply_text("Profilin sıfırlandı.")


# ============== PROFİL OLUŞTURMA (SORU-CEVAP AKIŞI) ==============
PROFIL_SORULARI = [
    ("yas", "Kaç yaşındasın?"),
    ("boy_kilo", "Boyun (cm) ve kilon (kg) nedir? (örn: 180 cm, 85 kg)"),
    ("hedef", "Ana hedefin ne? (örn: kilo vermek, kas kazanmak, bir yarışa hazırlanmak, genel fitness)"),
    ("sakatlik", "Herhangi bir sakatlığın ya da fiziksel kısıtlaman var mı? Yoksa 'yok' yaz."),
    ("deneyim", "Spor deneyim seviyeni nasıl tanımlarsın? (yeni başlayan / orta seviye / ileri seviye)"),
    ("siklik", "Haftada kaç gün antrenman yapabiliyorsun, hangi günler müsaitsin?"),
    ("ozel_hedef", "Ulaşmak istediğin özel bir hedef ya da yarış/tarih var mı? Yoksa 'yok' yaz."),
]


async def profil_olustur(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["profil_onboarding_index"] = 0
    context.user_data["profil_onboarding_cevaplar"] = {}
    await update.message.reply_text(
        "Harika, sana en iyi koçluğu yapabilmem için birkaç soru soracağım. "
        "İstediğin an /iptal yazarak durdurabilirsin.\n\n"
        f"1/{len(PROFIL_SORULARI)}: {PROFIL_SORULARI[0][1]}"
    )


async def _onboarding_cevabini_isle(update: Update, context: ContextTypes.DEFAULT_TYPE):
    index = context.user_data.get("profil_onboarding_index", 0)
    anahtar, soru_metni = PROFIL_SORULARI[index]
    context.user_data["profil_onboarding_cevaplar"][anahtar] = update.message.text

    yeni_index = index + 1
    if yeni_index < len(PROFIL_SORULARI):
        context.user_data["profil_onboarding_index"] = yeni_index
        await update.message.reply_text(
            f"{yeni_index + 1}/{len(PROFIL_SORULARI)}: {PROFIL_SORULARI[yeni_index][1]}"
        )
        return

    # Tüm sorular bitti — profili derleyip kaydet
    cevaplar = context.user_data["profil_onboarding_cevaplar"]
    kullanici_id = update.effective_user.id
    profil_metni = (
        f"- Yaş: {cevaplar.get('yas', '')}\n"
        f"- Boy/Kilo: {cevaplar.get('boy_kilo', '')}\n"
        f"- Ana hedef: {cevaplar.get('hedef', '')}\n"
        f"- Sakatlık/kısıtlama: {cevaplar.get('sakatlik', '')}\n"
        f"- Deneyim seviyesi: {cevaplar.get('deneyim', '')}\n"
        f"- Antrenman sıklığı: {cevaplar.get('siklik', '')}\n"
        f"- Özel hedef/yarış: {cevaplar.get('ozel_hedef', '')}"
    )
    profili_yaz(kullanici_id, profil_metni)

    context.user_data.pop("profil_onboarding_index", None)
    context.user_data.pop("profil_onboarding_cevaplar", None)

    await update.message.reply_text(
        f"✅ Profilin oluşturuldu, bundan sonra hep hatırlayacağım:\n\n{profil_metni}\n\n"
        f"İstediğin zaman /profil_ekle ile yeni bilgi ekleyebilir, /profil_goster ile "
        f"görebilir, /profil_olustur ile baştan yapabilirsin."
    )


async def profil_iptal(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if "profil_onboarding_index" in context.user_data:
        context.user_data.pop("profil_onboarding_index", None)
        context.user_data.pop("profil_onboarding_cevaplar", None)
        await update.message.reply_text("Profil oluşturma iptal edildi.")
    elif "olcum_onboarding_index" in context.user_data:
        context.user_data.pop("olcum_onboarding_index", None)
        context.user_data.pop("olcum_onboarding_cevaplar", None)
        await update.message.reply_text("Ölçüm ekleme iptal edildi.")
    else:
        await update.message.reply_text("İptal edilecek aktif bir işlem yok.")


# ============== VÜCUT ÖLÇÜMLERİ (AYLIK TAKİP) ==============
OLCUM_SORULARI = [
    ("kilo", "⚖️ Kaç kg'sın? (En doğru sonuç için sabah, aç karnına, tuvaletten sonra tart)"),
    ("kol", "💪 Kol (pazı) çevren kaç cm? Kolun gevşek/sarkık haldeyken, pazının EN KALIN noktasından mezura ile ölç."),
    ("omuz", "🧍 Omuz genişliğin kaç cm? Bir omuz ucundan diğer omuz ucuna, en geniş noktadan ölç."),
    ("gogus", "📏 Göğüs çevren kaç cm? Mezurayı meme ucu hizasından, göğsün en geniş noktasından geçirerek ölç, normal nefes alırken."),
    ("bel", "📐 Bel çevren kaç cm? Göbek deliği hizasından, nefesini verip rahat dururken ölç (içeri çekmeden)."),
    ("kalca", "🍑 Kalça çevren kaç cm? Kalçanın EN GENİŞ (en çıkıntılı) noktasından ölç."),
    ("bacak", "🦵 Uyluk (üst bacak) çevren kaç cm? Kasığın hemen altından, bacağın en kalın noktasından ölç."),
]


def olcumu_kaydet(kullanici_id, cevaplar):
    if not DATABASE_URL:
        return
    baglanti = psycopg2.connect(DATABASE_URL)
    baglanti.autocommit = True
    imlec = baglanti.cursor()
    imlec.execute(
        "INSERT INTO vucut_olculeri (kullanici_id, kilo, kol, omuz, gogus, bel, kalca, bacak) "
        "VALUES (%s, %s, %s, %s, %s, %s, %s, %s)",
        (kullanici_id, cevaplar.get("kilo"), cevaplar.get("kol"), cevaplar.get("omuz"),
         cevaplar.get("gogus"), cevaplar.get("bel"), cevaplar.get("kalca"), cevaplar.get("bacak")),
    )
    imlec.close()
    baglanti.close()


def son_iki_olcumu_getir(kullanici_id):
    if not DATABASE_URL:
        return []
    baglanti = psycopg2.connect(DATABASE_URL)
    imlec = baglanti.cursor()
    imlec.execute(
        "SELECT tarih, kilo, kol, omuz, gogus, bel, kalca, bacak FROM vucut_olculeri "
        "WHERE kullanici_id = %s ORDER BY tarih DESC LIMIT 2",
        (kullanici_id,),
    )
    satirlar = imlec.fetchall()
    imlec.close()
    baglanti.close()
    return satirlar


def tum_olcumleri_getir(kullanici_id):
    if not DATABASE_URL:
        return []
    baglanti = psycopg2.connect(DATABASE_URL)
    imlec = baglanti.cursor()
    imlec.execute(
        "SELECT tarih, kilo, kol, omuz, gogus, bel, kalca, bacak FROM vucut_olculeri "
        "WHERE kullanici_id = %s ORDER BY tarih ASC",
        (kullanici_id,),
    )
    satirlar = imlec.fetchall()
    imlec.close()
    baglanti.close()
    return satirlar


def olcum_ozeti_ve_trend(kullanici_id):
    """Son ölçüm ve varsa bir önceki ölçümle karşılaştırmalı değişimi
    metin olarak döndürür — antrenman/beslenme önerilerine otomatik
    dahil edilmek üzere."""
    son_ikisi = son_iki_olcumu_getir(kullanici_id)
    if not son_ikisi:
        return ""

    alanlar = ["kilo", "kol", "omuz", "gogus", "bel", "kalca", "bacak"]
    son = son_ikisi[0]
    satirlar = [f"En son ölçümlerim ({son[0].strftime('%d %B %Y') if hasattr(son[0], 'strftime') else son[0]}):"]
    for i, ad in enumerate(alanlar, start=1):
        if son[i]:
            satirlar.append(f"- {ad.capitalize()}: {son[i]}")

    if len(son_ikisi) > 1:
        onceki = son_ikisi[1]
        satirlar.append("\nBir önceki ölçüme göre değişim:")
        for i, ad in enumerate(alanlar, start=1):
            try:
                yeni_deger = float(str(son[i]).replace(",", "."))
                eski_deger = float(str(onceki[i]).replace(",", "."))
                fark = yeni_deger - eski_deger
                isaret = "+" if fark >= 0 else ""
                satirlar.append(f"- {ad.capitalize()}: {isaret}{fark:.1f}")
            except (ValueError, TypeError):
                continue

    return "\n".join(satirlar)


async def olcum_ekle(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["olcum_onboarding_index"] = 0
    context.user_data["olcum_onboarding_cevaplar"] = {}
    await update.message.reply_text(
        "Ölçümlerini alalım, aylık takip edip sana göre antrenman/beslenmeni "
        "ayarlayayım. İstersen /iptal ile durdurabilirsin.\n\n"
        f"1/{len(OLCUM_SORULARI)}: {OLCUM_SORULARI[0][1]}"
    )


async def _olcum_cevabini_isle(update: Update, context: ContextTypes.DEFAULT_TYPE):
    index = context.user_data.get("olcum_onboarding_index", 0)
    anahtar, _ = OLCUM_SORULARI[index]
    context.user_data["olcum_onboarding_cevaplar"][anahtar] = update.message.text

    yeni_index = index + 1
    if yeni_index < len(OLCUM_SORULARI):
        context.user_data["olcum_onboarding_index"] = yeni_index
        await update.message.reply_text(
            f"{yeni_index + 1}/{len(OLCUM_SORULARI)}: {OLCUM_SORULARI[yeni_index][1]}"
        )
        return

    kullanici_id = update.effective_user.id
    cevaplar = context.user_data["olcum_onboarding_cevaplar"]
    olcumu_kaydet(kullanici_id, cevaplar)

    context.user_data.pop("olcum_onboarding_index", None)
    context.user_data.pop("olcum_onboarding_cevaplar", None)

    await update.message.reply_text("✅ Ölçümlerin kaydedildi, yorumluyorum...")

    ozet = olcum_ozeti_ve_trend(kullanici_id)
    client_gemini, koleksiyon = istemcileri_al()
    yumusak = yumusak_ton_mu(kullanici_id)
    profil = profili_oku(kullanici_id)
    soru = (f"Yeni vücut ölçümlerimi ekledim, bunları yorumlar mısın ve gerekirse "
            f"antrenman/beslenme planımda ne değiştirmemi önerirsin?\n\n{ozet}")
    bulunan = await asyncio.to_thread(koleksiyon.query, query_texts=[soru], n_results=KAC_PARCA_GETIRILSIN)
    baglam, _ = baglami_hazirla(bulunan) if bulunan['documents'][0] else ("", [])
    gecmis = gecmisi_oku(kullanici_id)
    cevap = await asyncio.to_thread(
        cevap_uret, client_gemini, soru, baglam, gecmis, yumusak=yumusak, profil=profil
    )

    mesaji_kaydet(kullanici_id, "user", soru)
    mesaji_kaydet(kullanici_id, "model", cevap)
    await guvenli_reply(update.message, cevap)
async def olcum_gecmisi(update: Update, context: ContextTypes.DEFAULT_TYPE):
    kullanici_id = update.effective_user.id
    tumu = tum_olcumleri_getir(kullanici_id)
    if not tumu:
        await update.message.reply_text("Henüz hiç ölçüm eklemedin. /olcum_ekle ile başla.")
        return

    satirlar = ["📏 Ölçüm geçmişin:\n"]
    for satir in tumu:
        tarih = satir[0].strftime("%d %B %Y") if hasattr(satir[0], "strftime") else satir[0]
        satirlar.append(
            f"{tarih}: kilo {satir[1]}, kol {satir[2]}, omuz {satir[3]}, "
            f"göğüs {satir[4]}, bel {satir[5]}, kalça {satir[6]}, bacak {satir[7]}"
        )
    await guvenli_reply(update.message, "\n".join(satirlar))


async def olcum_hatirlatma_ac(update: Update, context: ContextTypes.DEFAULT_TYPE):
    kullanici_id = update.effective_user.id
    if DATABASE_URL:
        baglanti = psycopg2.connect(DATABASE_URL)
        baglanti.autocommit = True
        imlec = baglanti.cursor()
        imlec.execute(
            "INSERT INTO tg_ayarlar (kullanici_id, olcum_hatirlatma) VALUES (%s, TRUE) "
            "ON CONFLICT (kullanici_id) DO UPDATE SET olcum_hatirlatma = TRUE",
            (kullanici_id,),
        )
        imlec.close()
        baglanti.close()
    await update.message.reply_text("Ayda bir ölçüm hatırlatması göndereceğim.")


async def olcum_hatirlatma_kapat(update: Update, context: ContextTypes.DEFAULT_TYPE):
    kullanici_id = update.effective_user.id
    if DATABASE_URL:
        baglanti = psycopg2.connect(DATABASE_URL)
        baglanti.autocommit = True
        imlec = baglanti.cursor()
        imlec.execute(
            "INSERT INTO tg_ayarlar (kullanici_id, olcum_hatirlatma) VALUES (%s, FALSE) "
            "ON CONFLICT (kullanici_id) DO UPDATE SET olcum_hatirlatma = FALSE",
            (kullanici_id,),
        )
        imlec.close()
        baglanti.close()
    await update.message.reply_text("Ölçüm hatırlatmaları kapatıldı.")


async def olcum_hatirlatma_isi(context: ContextTypes.DEFAULT_TYPE):
    """Her gün kontrol eder: hatırlatması açık VE son ölçümünden 30+ gün
    geçmiş (ya da hiç ölçümü olmayan) kullanıcılara hatırlatma gönderir."""
    if not DATABASE_URL:
        return
    try:
        baglanti = psycopg2.connect(DATABASE_URL)
        imlec = baglanti.cursor()
        imlec.execute(
            "SELECT kullanici_id FROM tg_ayarlar WHERE olcum_hatirlatma = TRUE"
        )
        kullanicilar = [r[0] for r in imlec.fetchall()]
        imlec.close()
        baglanti.close()
    except Exception:
        return

    for kullanici_id in kullanicilar:
        try:
            son_ikisi = son_iki_olcumu_getir(kullanici_id)
            gonder = False
            if not son_ikisi:
                gonder = True
            else:
                son_tarih = son_ikisi[0][0]
                if hasattr(son_tarih, "date"):
                    gecen_gun = (datetime.now() - son_tarih.replace(tzinfo=None)).days
                    if gecen_gun >= 30:
                        gonder = True
            if gonder:
                await context.bot.send_message(
                    chat_id=kullanici_id,
                    text="📏 Aylık ölçüm zamanı geldi gibi! Vücut ölçülerini "
                         "güncellemek için /olcum_ekle yazabilirsin.",
                )
        except Exception as e:
            print(f"Ölçüm hatırlatma hatası (kullanıcı {kullanici_id}): {e}")


async def sabah_ac(update: Update, context: ContextTypes.DEFAULT_TYPE):
    kullanici_id = update.effective_user.id
    if DATABASE_URL:
        baglanti = psycopg2.connect(DATABASE_URL)
        baglanti.autocommit = True
        imlec = baglanti.cursor()
        imlec.execute(
            "INSERT INTO tg_ayarlar (kullanici_id, sabah_mesaji) VALUES (%s, TRUE) "
            "ON CONFLICT (kullanici_id) DO UPDATE SET sabah_mesaji = TRUE",
            (kullanici_id,),
        )
        imlec.close()
        baglanti.close()
    await update.message.reply_text("☀️ Tamamdır, her sabah 07:00'de sana otomatik bir mesaj göndereceğim.")


async def sabah_kapat(update: Update, context: ContextTypes.DEFAULT_TYPE):
    kullanici_id = update.effective_user.id
    if DATABASE_URL:
        baglanti = psycopg2.connect(DATABASE_URL)
        baglanti.autocommit = True
        imlec = baglanti.cursor()
        imlec.execute(
            "INSERT INTO tg_ayarlar (kullanici_id, sabah_mesaji) VALUES (%s, FALSE) "
            "ON CONFLICT (kullanici_id) DO UPDATE SET sabah_mesaji = FALSE",
            (kullanici_id,),
        )
        imlec.close()
        baglanti.close()
    await update.message.reply_text("Sabah mesajları kapatıldı.")


def sesli_cevap_mi(kullanici_id):
    if not DATABASE_URL:
        return False
    try:
        baglanti = psycopg2.connect(DATABASE_URL)
        imlec = baglanti.cursor()
        imlec.execute("SELECT sesli_cevap FROM tg_ayarlar WHERE kullanici_id = %s", (kullanici_id,))
        satir = imlec.fetchone()
        imlec.close()
        baglanti.close()
        return bool(satir[0]) if satir else False
    except Exception:
        return False


async def sesli_cevap_ac(update: Update, context: ContextTypes.DEFAULT_TYPE):
    kullanici_id = update.effective_user.id
    if DATABASE_URL:
        baglanti = psycopg2.connect(DATABASE_URL)
        baglanti.autocommit = True
        imlec = baglanti.cursor()
        imlec.execute(
            "INSERT INTO tg_ayarlar (kullanici_id, sesli_cevap) VALUES (%s, TRUE) "
            "ON CONFLICT (kullanici_id) DO UPDATE SET sesli_cevap = TRUE",
            (kullanici_id,),
        )
        imlec.close()
        baglanti.close()
    await update.message.reply_text(
        "🔊 Tamamdır, bundan sonra cevaplarımı hem yazı hem sesli mesaj olarak "
        "göndereceğim — spor salonunda eller boşken de dinleyebilirsin."
    )


async def sesli_cevap_kapat(update: Update, context: ContextTypes.DEFAULT_TYPE):
    kullanici_id = update.effective_user.id
    if DATABASE_URL:
        baglanti = psycopg2.connect(DATABASE_URL)
        baglanti.autocommit = True
        imlec = baglanti.cursor()
        imlec.execute(
            "INSERT INTO tg_ayarlar (kullanici_id, sesli_cevap) VALUES (%s, FALSE) "
            "ON CONFLICT (kullanici_id) DO UPDATE SET sesli_cevap = FALSE",
            (kullanici_id,),
        )
        imlec.close()
        baglanti.close()
    await update.message.reply_text("Sesli cevaplar kapatıldı, sadece yazı ile devam.")


async def intervals_baglan(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Kullanım: /intervals_baglan <api_key>
    Intervals.icu → Settings → Developer Settings'ten alınan API anahtarı."""
    if not context.args:
        await update.message.reply_text(
            "Kullanım: /intervals_baglan <api_key>\n\n"
            "intervals.icu → Settings → Developer Settings kısmından "
            "API anahtarını alıp buraya yapıştır."
        )
        return

    api_key = context.args[0]
    kullanici_id = update.effective_user.id

    try:
        veri = await asyncio.to_thread(intervals_uyku_verisi_getir, api_key)
        intervals_baglantisini_kaydet(kullanici_id, api_key)

        # Bağlanma anında var olan aktiviteleri baştan "görülmüş" işaretle
        # — yoksa geçmiş antrenmanlar "yeni" diye bildirilmeye başlar.
        try:
            mevcut_aktiviteler = await asyncio.to_thread(intervals_aktiviteleri_getir, api_key, 3)
            for a in mevcut_aktiviteler:
                aid = str(a.get("id", ""))
                if aid:
                    intervals_aktiviteyi_bildirildi_isaretle(kullanici_id, aid)
        except Exception:
            pass

        await update.message.reply_text(
            "✅ Intervals.icu bağlandı! Artık uyku/HRV verilerini "
            "sabah mesajlarında ve isteğinde kullanabileceğim."
        )
    except Exception as e:
        await update.message.reply_text(f"Bağlantı başarısız: {e}")


async def uyku_durumu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    kullanici_id = update.effective_user.id
    api_key = intervals_api_key_getir(kullanici_id)
    if not api_key:
        await update.message.reply_text(
            "Intervals.icu hesabın bağlı değil. Bağlamak için /intervals_baglan yaz."
        )
        return

    try:
        veri = await asyncio.to_thread(intervals_uyku_verisi_getir, api_key)
        ozet = intervals_ozet_metni(veri)
        if not ozet:
            await update.message.reply_text("Bugün için henüz uyku/HRV verisi gelmemiş.")
            return

        client_gemini, koleksiyon = istemcileri_al()
        yumusak = yumusak_ton_mu(kullanici_id)
        soru = (
            f"[NOT: Bu, Intervals.icu'dan ŞU AN çektiğim EN GÜNCEL uyku "
            f"verisi — gece boyunca uyku güncellenmiş olabilir (örn. "
            f"erken uyanıp tekrar uyumuş olabilirim). Eğer bu rakam, "
            f"bugün daha önce konuştuğumuz bir uyku bilgisinden farklıysa, "
            f"bu bir ÇELİŞKİ DEĞİL — sadece veri güncellenmiş demektir, "
            f"ÖNCEKİ değil BU en güncel veriye güven, farklılığı "
            f"'tutarsızlık' diye sorgulama.]\n\n"
            f"Bugünkü uyku/toparlanma verilerimi yorumlar mısın, "
            f"antrenmanımı buna göre ayarlamalı mıyım?\n\n{ozet}"
        )
        bulunan = await asyncio.to_thread(koleksiyon.query, query_texts=[soru], n_results=KAC_PARCA_GETIRILSIN)
        baglam, _ = baglami_hazirla(bulunan) if bulunan['documents'][0] else ("", [])
        gecmis = gecmisi_oku(kullanici_id)
        profil = profili_oku(kullanici_id)
        cevap = await asyncio.to_thread(
            cevap_uret, client_gemini, soru, baglam, gecmis, yumusak=yumusak, profil=profil
        )

        mesaji_kaydet(kullanici_id, "user", soru)
        mesaji_kaydet(kullanici_id, "model", cevap)
        await guvenli_reply(update.message, cevap)
    except Exception as e:
        await update.message.reply_text(f"Uyku verisi alınırken hata: {e}")


async def antrenman_gecmisi(update: Update, context: ContextTypes.DEFAULT_TYPE):
    kullanici_id = update.effective_user.id
    gun_sayisi = 14
    if context.args:
        try:
            gun_sayisi = int(context.args[0])
        except ValueError:
            pass
    ozet = antrenman_gunlugu_ozeti(kullanici_id, gun_sayisi=gun_sayisi)
    if not ozet:
        await update.message.reply_text("Bu dönemde kayıtlı bir antrenman günlüğü bulamadım.")
        return
    await guvenli_reply(update.message, "📓 " + ozet)


async def antrenman_ekle(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Otomatik tespitin kaçırdığı durumlar için GARANTİLİ manuel kayıt.
    Kullanım: /antrenman_ekle Bacak günü: Squat 5x5 100kg, RDL 3x8"""
    if not context.args:
        await update.message.reply_text(
            "Kullanım: /antrenman_ekle <açıklama>\n"
            "Örn: /antrenman_ekle Bacak günü: Squat 5x5 100kg, RDL 3x8 60kg"
        )
        return
    kullanici_id = update.effective_user.id
    aciklama = " ".join(context.args)
    antrenman_kaydet(kullanici_id, aciklama)
    await update.message.reply_text("✅ Antrenman günlüğüne eklendi, garanti şekilde hatırlayacağım.")


# ============== BESLENME TAKİBİ (FOTOĞRAFLI) ==============
def beslenme_kaydet(kullanici_id, aciklama, kalori, protein, karbonhidrat, yag):
    if not DATABASE_URL:
        return
    baglanti = psycopg2.connect(DATABASE_URL)
    baglanti.autocommit = True
    imlec = baglanti.cursor()
    imlec.execute(
        "INSERT INTO beslenme_kayitlari (kullanici_id, aciklama, tahmini_kalori, "
        "tahmini_protein, tahmini_karbonhidrat, tahmini_yag) VALUES (%s, %s, %s, %s, %s, %s)",
        (kullanici_id, aciklama, kalori, protein, karbonhidrat, yag),
    )
    imlec.close()
    baglanti.close()


def beslenme_ozeti_getir(kullanici_id, gun_sayisi=1):
    """gun_sayisi=1 -> SADECE bugün (Türkiye takvim günü), 'son 24 saat'
    değil. Eskiden 'son 24 saat' penceresi kullanılıyordu, bu da dünün
    geç saatlerindeki yemeklerin 'bugünmüş' gibi görünmesine sebep
    oluyordu — artık gerçek takvim gününe göre filtreleniyor."""
    if not DATABASE_URL:
        return []
    baglanti = psycopg2.connect(DATABASE_URL)
    imlec = baglanti.cursor()
    imlec.execute(
        "SELECT tarih, aciklama, tahmini_kalori, tahmini_protein, tahmini_karbonhidrat, "
        "tahmini_yag FROM beslenme_kayitlari WHERE kullanici_id = %s AND "
        "(tarih AT TIME ZONE 'UTC' AT TIME ZONE 'Europe/Istanbul')::date >= "
        "(NOW() AT TIME ZONE 'Europe/Istanbul')::date - %s::int "
        "ORDER BY tarih ASC",
        (kullanici_id, gun_sayisi - 1),
    )
    satirlar = imlec.fetchall()
    imlec.close()
    baglanti.close()
    return satirlar


async def yemek_ekle(update: Update, context: ContextTypes.DEFAULT_TYPE):
    su_an_acik = context.user_data.get("yemek_modu", False)
    if su_an_acik:
        context.user_data["yemek_modu"] = False
        await update.message.reply_text("Yemek modu kapatıldı, fotoğraflar artık normal soru gibi işlenecek.")
    else:
        context.user_data["yemek_modu"] = True
        await update.message.reply_text(
            "📸 Yemek modu açık — bundan sonra gönderdiğin HER fotoğrafı otomatik "
            "yemek olarak kaydedip kalori/makro tahmini yapacağım. Kapatmak için "
            "tekrar /yemek_ekle yaz."
        )


async def _yemek_fotografini_isle(update: Update, context: ContextTypes.DEFAULT_TYPE, foto_bytes):
    client_gemini, _ = istemcileri_al()
    gorsel_b64 = base64.b64encode(foto_bytes).decode("utf-8")
    kullanici_id = update.effective_user.id
    kullanici_notu = update.message.caption or ""

    talimat = (
        "Bu fotoğraftaki yemeği/öğünü incele. SADECE aşağıdaki JSON formatında "
        "cevap ver, başka hiçbir şey yazma:\n"
        '{"aciklama": "kısa açıklama", "kalori": 000, "protein": 00, '
        '"karbonhidrat": 00, "yag": 00}\n'
        "Değerler tahmini olabilir ama makul olsun (gram/kalori cinsinden sayı, birim yazma). "
        + (
            f"\n\nÖNEMLİ: Kullanıcı bu fotoğrafa şu notu eklemiş, bunu MUTLAKA dikkate al "
            f"(porsiyon/malzeme bilgisi verdiyse tahminini buna göre kesinleştir): "
            f"\"{kullanici_notu}\""
            if kullanici_notu else
            "\n\nKullanıcı porsiyon/malzeme belirtmedi, görselden en makul tahmini yap "
            "ama bunun kaba bir tahmin olduğunu unutma."
        )
    )
    veri = None
    son_hata = None
    for model_adi in GEMINI_MODEL_LISTESI:
        try:
            yanit = await asyncio.to_thread(
                client_gemini.models.generate_content,
                model=model_adi,
                contents=[
                    {"role": "user", "parts": [
                        {"text": talimat},
                        {"inline_data": {"mime_type": "image/jpeg", "data": gorsel_b64}},
                    ]}
                ],
            )
            metin = re.sub(r"^```json\s*|\s*```$", "", yanit.text.strip(), flags=re.MULTILINE).strip("`").strip()
            veri = json.loads(metin)
            break
        except Exception as e:
            son_hata = e
            continue

    if veri is None:
        await update.message.reply_text(f"Yemek analiz edilemedi (tüm modeller yoğun): {son_hata}")
        return

    beslenme_kaydet(
        kullanici_id, veri.get("aciklama", ""), veri.get("kalori"),
        veri.get("protein"), veri.get("karbonhidrat"), veri.get("yag"),
    )

    await update.message.reply_text(
        f"✅ Kaydedildi: {veri.get('aciklama', '')}\n"
        f"~{veri.get('kalori', '?')} kcal | "
        f"P: {veri.get('protein', '?')}g | "
        f"K: {veri.get('karbonhidrat', '?')}g | "
        f"Y: {veri.get('yag', '?')}g"
    )

    # Kayıttan sonra, kullanıcının profiline/hedefine göre kısa bir
    # koçluk yorumu da ekle — sadece kayıt değil, gerçek değerlendirme.
    try:
        yumusak = yumusak_ton_mu(kullanici_id)
        profil = profili_oku(kullanici_id)
        olcum_ozeti = olcum_ozeti_ve_trend(kullanici_id)
        if olcum_ozeti:
            profil = (profil + "\n\n" + olcum_ozeti).strip() if profil else olcum_ozeti
        gunluk_beslenme = beslenme_gunluk_ozet_metni(kullanici_id)
        if gunluk_beslenme:
            profil = (profil + "\n\n" + gunluk_beslenme).strip() if profil else gunluk_beslenme

        soru = (
            f"[FOTOĞRAF ANALİZİ — BEN (antrenör) fotoğrafa bakarak ÜRETTİM, "
            f"kullanıcı bu sayıları bana YAZMADI/SÖYLEMEDİ, benim görsel "
            f"tahminimdir] Öğün: {veri.get('aciklama', '')} "
            f"(~{veri.get('kalori', '?')} kcal, P:{veri.get('protein', '?')}g, "
            f"K:{veri.get('karbonhidrat', '?')}g, Y:{veri.get('yag', '?')}g). "
            f"Kısaca yorumlar mısın — hedefime uygun mu, bir sonraki öğünde "
            f"nelere dikkat etmeliyim?"
        )
        _, koleksiyon = istemcileri_al()
        bulunan = await asyncio.to_thread(koleksiyon.query, query_texts=[soru], n_results=KAC_PARCA_GETIRILSIN)
        baglam, _ = baglami_hazirla(bulunan) if bulunan['documents'][0] else ("", [])
        gecmis = gecmisi_oku(kullanici_id)
        yorum = await asyncio.to_thread(
            cevap_uret, client_gemini, soru, baglam, gecmis, yumusak=yumusak, profil=profil
        )

        mesaji_kaydet(kullanici_id, "user", soru)
        mesaji_kaydet(kullanici_id, "model", yorum)
        await guvenli_reply(update.message, yorum)
    except Exception as e:
        print(f"Yemek yorumu üretilirken hata: {e}")


async def yemek_duzelt(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Kullanım: /yemek_duzelt <kalori> <protein> <karbonhidrat> <yag>
    En son eklediğin yemek kaydının değerlerini, gerçek (örn. paket
    etiketinden bildiğin) rakamlarla düzeltir."""
    if len(context.args) < 4:
        await update.message.reply_text(
            "Kullanım: /yemek_duzelt <kalori> <protein> <karbonhidrat> <yag>\n"
            "Örn: /yemek_duzelt 450 30 40 15"
        )
        return

    try:
        kalori, protein, karbonhidrat, yag = [int(x) for x in context.args[:4]]
    except ValueError:
        await update.message.reply_text("Değerler sayı olmalı. Örn: /yemek_duzelt 450 30 40 15")
        return

    kullanici_id = update.effective_user.id
    if not DATABASE_URL:
        await update.message.reply_text("Veritabanı bağlantısı yok.")
        return

    baglanti = psycopg2.connect(DATABASE_URL)
    baglanti.autocommit = True
    imlec = baglanti.cursor()
    imlec.execute(
        "SELECT id FROM beslenme_kayitlari WHERE kullanici_id = %s "
        "ORDER BY tarih DESC LIMIT 1", (kullanici_id,),
    )
    satir = imlec.fetchone()
    if not satir:
        await update.message.reply_text("Düzeltilecek bir kaydın yok.")
        imlec.close()
        baglanti.close()
        return

    imlec.execute(
        "UPDATE beslenme_kayitlari SET tahmini_kalori = %s, tahmini_protein = %s, "
        "tahmini_karbonhidrat = %s, tahmini_yag = %s WHERE id = %s",
        (kalori, protein, karbonhidrat, yag, satir[0]),
    )
    imlec.close()
    baglanti.close()
    await update.message.reply_text(f"✅ Düzeltildi: {kalori} kcal | P:{protein}g | K:{karbonhidrat}g | Y:{yag}g")


async def beslenme_ozet(update: Update, context: ContextTypes.DEFAULT_TYPE):
    kullanici_id = update.effective_user.id
    gun_sayisi = 1
    if context.args:
        try:
            gun_sayisi = int(context.args[0])
        except ValueError:
            pass

    kayitlar = beslenme_ozeti_getir(kullanici_id, gun_sayisi)
    if not kayitlar:
        await update.message.reply_text(
            f"Son {gun_sayisi} günde hiç beslenme kaydın yok. /yemek_ekle ile ekleyebilirsin."
        )
        return

    toplam_kalori = sum((k[2] or 0) for k in kayitlar)
    toplam_protein = sum((k[3] or 0) for k in kayitlar)
    toplam_karbonhidrat = sum((k[4] or 0) for k in kayitlar)
    toplam_yag = sum((k[5] or 0) for k in kayitlar)

    satirlar = [f"📊 Son {gun_sayisi} gün beslenme özeti ({len(kayitlar)} kayıt):\n"]
    satirlar.append(f"Toplam: ~{toplam_kalori} kcal | P: {toplam_protein}g | "
                     f"K: {toplam_karbonhidrat}g | Y: {toplam_yag}g")
    if gun_sayisi > 1:
        satirlar.append(f"Günlük ortalama: ~{toplam_kalori // gun_sayisi} kcal")
    satirlar.append("\nKayıtlar:")
    for tarih, aciklama, kalori, _, _, _ in kayitlar:
        tarih_str = tarih.strftime("%d.%m %H:%M") if hasattr(tarih, "strftime") else str(tarih)
        satirlar.append(f"- {tarih_str}: {aciklama} (~{kalori} kcal)")

    await guvenli_reply(update.message, "\n".join(satirlar))


def beslenme_gunluk_ozet_metni(kullanici_id):
    """cevap_uret'e otomatik verilecek bugünkü beslenme özeti — her
    öğünü GERÇEK saatiyle (Türkiye saatine çevrilmiş) ve açıklamasıyla
    birlikte gösterir, sadece toplam sayı değil."""
    kayitlar = beslenme_ozeti_getir(kullanici_id, gun_sayisi=1)
    if not kayitlar:
        return ""
    satirlar = ["Bugün şu ana kadar yediklerim (gerçek, kayıtlı öğünler, saatleriyle):"]
    toplam_kalori = 0
    for tarih, aciklama, kalori, protein, karbonhidrat, yag in kayitlar:
        try:
            from zoneinfo import ZoneInfo
            if hasattr(tarih, "tzinfo"):
                tarih_tr = tarih.replace(tzinfo=ZoneInfo("UTC")).astimezone(ZoneInfo("Europe/Istanbul")) \
                    if tarih.tzinfo is None else tarih.astimezone(ZoneInfo("Europe/Istanbul"))
                saat_str = tarih_tr.strftime("%H:%M")
            else:
                saat_str = str(tarih)
        except Exception:
            saat_str = tarih.strftime("%H:%M") if hasattr(tarih, "strftime") else str(tarih)
        satirlar.append(f"- Saat {saat_str}: {aciklama} (~{kalori or 0} kcal)")
        toplam_kalori += (kalori or 0)
    satirlar.append(f"Toplam bugün: ~{toplam_kalori} kcal, {len(kayitlar)} öğün")
    return "\n".join(satirlar)



async def sabah_mesaji_isi(context: ContextTypes.DEFAULT_TYPE):
    """Sabit saatte DEĞİL — ama SADECE 06:00-12:00 (Türkiye saati)
    PENCERESİ İÇİNDE çalışır. Bu pencerede, o günün uyku verisi
    Intervals.icu'ya düşmüş mü diye bakar. Düştüyse mesajı GÖNDERİR ve
    o günü 'gönderildi' işaretler. Veri hâlâ yoksa ve saat 11:00'i
    geçtiyse (ama hâlâ 12:00'den önceyse), veri olmasa bile genel bir
    mesaj gönderir. PENCERENİN DIŞINDA (12:00-06:00 arası) HİÇBİR ZAMAN
    göndermez — bu, deploy/yeniden başlatmaların yanlışlıkla gece
    'Günaydın' mesajı göndermesini önlemek için kritik bir sınır."""
    if not DATABASE_URL:
        return

    su_an = _turkiye_simdi()
    if not (6 <= su_an.hour < 12):
        return  # pencere dışında — ne veri kontrolü ne fallback, hiç çalışma

    try:
        baglanti = psycopg2.connect(DATABASE_URL)
        imlec = baglanti.cursor()
        imlec.execute(
            "SELECT kullanici_id FROM tg_ayarlar WHERE sabah_mesaji = TRUE "
            "AND (son_sabah_mesaji_tarihi IS NULL OR son_sabah_mesaji_tarihi < %s)",
            (su_an.date(),),
        )
        kullanicilar = [r[0] for r in imlec.fetchall()]
        imlec.close()
        baglanti.close()
    except Exception:
        return

    if not kullanicilar:
        return  # bugün herkese zaten gönderilmiş

    son_tarih_gecerli = su_an.hour >= 11  # 11:00-12:00 arası: veri beklemeden gönder

    client_gemini, koleksiyon = istemcileri_al()

    for kullanici_id in kullanicilar:
        try:
            uyku_ozeti = ""
            intervals_key = intervals_api_key_getir(kullanici_id)
            veri_hazir = False
            if intervals_key:
                try:
                    uyku_verisi = await asyncio.to_thread(intervals_uyku_verisi_getir, intervals_key)
                    uyku_metni = intervals_ozet_metni(uyku_verisi)
                    if uyku_metni:
                        uyku_ozeti = f"\n\n{uyku_metni}"
                        veri_hazir = True
                except Exception:
                    pass

            # Uyku verisi yoksa VE henüz 11:00 olmadıysa, bu turu atla —
            # kullanıcı muhtemelen hâlâ uyuyor, 30 dk sonra tekrar bakılır.
            if not veri_hazir and not son_tarih_gecerli and intervals_key:
                continue

            yumusak = yumusak_ton_mu(kullanici_id)
            profil = profili_oku(kullanici_id)
            gecmis = gecmisi_oku(kullanici_id, limit=15)

            strava_ozeti = ""
            baglanti_bilgisi = strava_baglantisini_getir(kullanici_id)
            if baglanti_bilgisi:
                try:
                    access_token = await asyncio.to_thread(
                        strava_erisim_tokeni_al, baglanti_bilgisi["refresh_token"]
                    )
                    son_aktiviteler = await asyncio.to_thread(
                        strava_son_aktiviteleri_getir, access_token, kac_tane=1
                    )
                    if son_aktiviteler:
                        strava_ozeti = f"\n\nEn son aktivitem:\n{strava_aktiviteyi_metne_cevir(son_aktiviteler[0])}"
                except Exception:
                    pass

            soru = (f"Günaydın koç! Bugün için bana kısa bir motivasyon ve gün planı "
                    f"önerir misin?{strava_ozeti}{uyku_ozeti}")
            bulunan = await asyncio.to_thread(koleksiyon.query, query_texts=[soru], n_results=KAC_PARCA_GETIRILSIN)
            baglam, _ = baglami_hazirla(bulunan) if bulunan['documents'][0] else ("", [])
            cevap = await asyncio.to_thread(
                cevap_uret, client_gemini, soru, baglam, gecmis, yumusak=yumusak, profil=profil
            )

            mesaji_kaydet(kullanici_id, "user", soru)
            mesaji_kaydet(kullanici_id, "model", cevap)
            await guvenli_send_message(context.bot, kullanici_id, f"☀️ Günaydın!\n\n{cevap}")

            baglanti2 = psycopg2.connect(DATABASE_URL)
            baglanti2.autocommit = True
            imlec2 = baglanti2.cursor()
            imlec2.execute(
                "UPDATE tg_ayarlar SET son_sabah_mesaji_tarihi = %s WHERE kullanici_id = %s",
                (_turkiye_simdi().date(), kullanici_id),
            )
            imlec2.close()
            baglanti2.close()
        except Exception as e:
            print(f"Sabah mesajı hatası (kullanıcı {kullanici_id}): {e}")


async def web_sohbetlerini_getir(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Koçum'un web (Chainlit) sürümündeki sohbetleri, aynı PostgreSQL
    veritabanından okuyup Telegram'ın kendi basit tablosuna kopyalar.
    Kullanım: /web_sohbetlerini_getir seninmailin@gmail.com"""
    if not context.args:
        await update.message.reply_text(
            "Kullanım: /web_sohbetlerini_getir email@adresin.com\n"
            "(Koçum web'e giriş yaptığın e-posta neyse onu yaz)"
        )
        return

    web_email = context.args[0]
    kullanici_id = update.effective_user.id

    if not DATABASE_URL:
        await update.message.reply_text("Veritabanı bağlantısı yok, işlem yapılamadı.")
        return

    try:
        baglanti = psycopg2.connect(DATABASE_URL)
        imlec = baglanti.cursor()

        imlec.execute('SELECT id FROM users WHERE identifier = %s', (web_email,))
        satir = imlec.fetchone()
        if not satir:
            await update.message.reply_text(
                f"'{web_email}' ile web'de giriş yapılmış bir hesap bulamadım."
            )
            imlec.close()
            baglanti.close()
            return
        web_user_id = satir[0]

        imlec.execute(
            'SELECT id, name FROM threads WHERE "userId" = %s ORDER BY "createdAt" ASC',
            (web_user_id,),
        )
        threadler = imlec.fetchall()

        toplam_mesaj = 0
        for thread_id, thread_adi in threadler:
            imlec.execute(
                'SELECT type, output, input, "createdAt" FROM steps '
                'WHERE "threadId" = %s AND type IN (\'user_message\', \'assistant_message\') '
                'ORDER BY "createdAt" ASC',
                (thread_id,),
            )
            adimlar = imlec.fetchall()
            for tur, cikti, girdi, _zaman in adimlar:
                icerik = cikti or girdi or ""
                if not icerik.strip():
                    continue
                rol = "user" if tur == "user_message" else "model"
                mesaji_kaydet(kullanici_id, rol, icerik)
                toplam_mesaj += 1

        imlec.close()
        baglanti.close()

        await update.message.reply_text(
            f"✅ Web'deki {len(threadler)} sohbetten toplam {toplam_mesaj} mesaj "
            f"buraya (gerçek konuşma geçmişine) aktarıldı. Artık onları da hatırlıyorum."
        )
    except Exception as e:
        await update.message.reply_text(f"Aktarma sırasında hata oluştu: {e}")


async def strava_baglan(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Kullanım: /strava_baglan <refresh_token>
    Her kullanıcı KENDİ Strava hesabını, kendi refresh_token'ıyla bağlar —
    böylece hesaplar birbirine karışmaz, herkesin verisi kendine özel kalır."""
    if not context.args:
        await update.message.reply_text(
            "Kullanım: /strava_baglan <refresh_token>\n\n"
            "Kendi Strava hesabını bağlamak için:\n"
            "1) strava.com/settings/api adresinden bir uygulama oluştur\n"
            "2) Bana Client ID ve Secret'ı gönder, sana yetkilendirme linki hazırlayayım\n"
            "3) O linkten aldığın kodu bana ilet, refresh_token üreteyim\n"
            "4) O refresh_token'ı bu komutla kaydet"
        )
        return

    refresh_token = context.args[0]
    kullanici_id = update.effective_user.id

    try:
        access_token = await asyncio.to_thread(strava_erisim_tokeni_al, refresh_token)
        aktiviteler = await asyncio.to_thread(strava_son_aktiviteleri_getir, access_token, kac_tane=1)
        athlete_id = None
        strava_baglantisini_kaydet(kullanici_id, refresh_token, athlete_id)
        if aktiviteler:
            strava_son_gorulen_guncelle(kullanici_id, aktiviteler[0]["id"])
        await update.message.reply_text(
            "✅ Strava hesabın başarıyla bağlandı! Artık antrenmanlarını görebiliyorum. "
            "Yeni bir aktivite bitirdiğinde sana otomatik haber vereceğim."
        )
    except Exception as e:
        await update.message.reply_text(f"Bağlantı başarısız: {e}")


async def video_var_mi(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Kullanım: /video_var_mi <video_id_veya_link>
    Bir videonun gerçekten arşivde (ChromaDB'de) olup olmadığını,
    varsa kaç parçaya bölündüğünü ve içeriğinden bir önizlemeyi gösterir."""
    if not context.args:
        await update.message.reply_text("Kullanım: /video_var_mi <video_id_veya_link>")
        return

    girdi = context.args[0]
    video_id = girdi
    eslesme = re.search(r"watch\?v=([\w-]+)", girdi)
    if eslesme:
        video_id = eslesme.group(1)

    try:
        _, koleksiyon = istemcileri_al()
        sonuc = koleksiyon.get(where={"video_id": video_id})
        if not sonuc.get("ids"):
            await update.message.reply_text(f"❌ '{video_id}' arşivde bulunamadı.")
            return
        onizleme = sonuc["documents"][0][:300] if sonuc.get("documents") else ""
        await update.message.reply_text(
            f"✅ '{video_id}' arşivde var! ({len(sonuc['ids'])} parça)\n\n"
            f"İçerik önizlemesi:\n{onizleme}..."
        )
    except Exception as e:
        await update.message.reply_text(f"Kontrol sırasında hata: {e}")


async def son_antrenman(update: Update, context: ContextTypes.DEFAULT_TYPE):
    kullanici_id = update.effective_user.id
    strava_baglanti = strava_baglantisini_getir(kullanici_id)
    intervals_key = intervals_api_key_getir(kullanici_id)

    if not strava_baglanti and not intervals_key:
        await update.message.reply_text(
            "Hiçbir hesabın bağlı değil. /strava_baglan ya da /intervals_baglan ile bağla."
        )
        return

    adaylar = []  # (tarih_str, aktivite_metni)

    if strava_baglanti:
        try:
            access_token = await asyncio.to_thread(strava_erisim_tokeni_al, strava_baglanti["refresh_token"])
            aktiviteler = await asyncio.to_thread(strava_son_aktiviteleri_getir, access_token, kac_tane=1)
            if aktiviteler:
                adaylar.append((aktiviteler[0].get("start_date_local", ""), strava_aktiviteyi_metne_cevir(aktiviteler[0])))
        except Exception as e:
            print(f"Strava son_antrenman hatası: {e}")

    if intervals_key:
        try:
            aktiviteler = await asyncio.to_thread(intervals_aktiviteleri_getir, intervals_key, 7)
            if aktiviteler:
                en_son = sorted(aktiviteler, key=lambda a: a.get("start_date_local", ""), reverse=True)[0]
                adaylar.append((en_son.get("start_date_local", ""), intervals_aktiviteyi_metne_cevir(en_son)))
        except Exception as e:
            print(f"Intervals.icu son_antrenman hatası: {e}")

    if not adaylar:
        await update.message.reply_text("Henüz hiç aktivite bulamadım.")
        return

    # İki kaynaktan da veri geldiyse, tarihi en YAKIN (en güncel) olanı kullan
    adaylar.sort(key=lambda x: x[0], reverse=True)
    aktivite_metni = adaylar[0][1]

    try:
        client_gemini, koleksiyon = istemcileri_al()
        yumusak = yumusak_ton_mu(kullanici_id)

        soru = f"Az önce bitirdiğim antrenmanı yorumlar mısın?\n\n{aktivite_metni}"
        bulunan = await asyncio.to_thread(koleksiyon.query, query_texts=[soru], n_results=KAC_PARCA_GETIRILSIN)
        baglam, _ = baglami_hazirla(bulunan) if bulunan['documents'][0] else ("", [])
        gecmis = gecmisi_oku(kullanici_id)
        tam_profil = tam_profili_olustur(kullanici_id)
        cevap = await asyncio.to_thread(cevap_uret, client_gemini, soru, baglam, gecmis, yumusak=yumusak, profil=tam_profil)

        mesaji_kaydet(kullanici_id, "user", soru)
        mesaji_kaydet(kullanici_id, "model", cevap)
        antrenman_kaydet_tekrarsiz(kullanici_id, f"[/son_antrenman] {aktivite_metni}", adaylar[0][0])
        await guvenli_reply(update.message, cevap)
    except Exception as e:
        await update.message.reply_text(f"Antrenman getirilirken hata: {e}")


async def strava_ozet(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Kullanım: /strava_ozet [gun_sayisi] — varsayılan son 7 gün.
    Toplam mesafe, süre, aktivite sayısı gibi TREND bilgisini özetler."""
    kullanici_id = update.effective_user.id
    baglanti_bilgisi = strava_baglantisini_getir(kullanici_id)
    if not baglanti_bilgisi:
        await update.message.reply_text("Strava hesabın bağlı değil. Bağlamak için /strava_baglan yaz.")
        return

    gun_sayisi = 7
    if context.args:
        try:
            gun_sayisi = int(context.args[0])
        except ValueError:
            pass

    try:
        access_token = await asyncio.to_thread(
            strava_erisim_tokeni_al, baglanti_bilgisi["refresh_token"]
        )
        aktiviteler = await asyncio.to_thread(strava_son_aktiviteleri_getir, access_token, kac_tane=100)

        esik_tarih = datetime.now() - timedelta(days=gun_sayisi)
        secilenler = []
        for a in aktiviteler:
            try:
                a_tarih = datetime.fromisoformat(a["start_date_local"].replace("Z", ""))
                if a_tarih >= esik_tarih:
                    secilenler.append(a)
            except Exception:
                continue

        if not secilenler:
            await update.message.reply_text(f"Son {gun_sayisi} günde hiç aktivite bulamadım.")
            return

        toplam_mesafe = sum((a.get("distance", 0) or 0) for a in secilenler) / 1000
        toplam_sure = sum((a.get("moving_time", 0) or 0) for a in secilenler) / 60
        toplam_tirmanis = sum((a.get("total_elevation_gain", 0) or 0) for a in secilenler)
        tur_sayisi = {}
        for a in secilenler:
            tur_sayisi[a.get("type", "?")] = tur_sayisi.get(a.get("type", "?"), 0) + 1

        ozet_metni = (
            f"Son {gun_sayisi} gün özeti:\n"
            f"- Toplam aktivite: {len(secilenler)}\n"
            f"- Toplam mesafe: {toplam_mesafe:.1f} km\n"
            f"- Toplam süre: {toplam_sure:.0f} dakika\n"
            f"- Toplam tırmanış: {toplam_tirmanis:.0f} m\n"
            f"- Türlere göre: {tur_sayisi}"
        )

        client_gemini, koleksiyon = istemcileri_al()
        yumusak = yumusak_ton_mu(kullanici_id)
        soru = f"Son {gun_sayisi} günlük antrenman verilerimi değerlendirir misin, trend olarak nasılım?\n\n{ozet_metni}"
        bulunan = await asyncio.to_thread(koleksiyon.query, query_texts=[soru], n_results=KAC_PARCA_GETIRILSIN)
        baglam, _ = baglami_hazirla(bulunan) if bulunan['documents'][0] else ("", [])
        gecmis = gecmisi_oku(kullanici_id)
        profil = profili_oku(kullanici_id)
        cevap = await asyncio.to_thread(
            cevap_uret, client_gemini, soru, baglam, gecmis, yumusak=yumusak, profil=profil
        )

        mesaji_kaydet(kullanici_id, "user", soru)
        mesaji_kaydet(kullanici_id, "model", cevap)
        await guvenli_reply(update.message, cevap)
    except Exception as e:
        await update.message.reply_text(f"Özet oluşturulurken hata: {e}")


async def strava_kontrol_isi(context: ContextTypes.DEFAULT_TYPE):
    """Periyodik olarak (JobQueue ile) her bağlı kullanıcının yeni bir
    Strava aktivitesi olup olmadığını kontrol eder, varsa OTOMATİK
    olarak SADECE O KULLANICIYA yorum gönderir — başkasına gitmez."""
    if not DATABASE_URL:
        return
    try:
        baglanti = psycopg2.connect(DATABASE_URL)
        imlec = baglanti.cursor()
        imlec.execute("SELECT kullanici_id, refresh_token, son_gorulen_aktivite_id FROM strava_baglantilar")
        tum_baglantilar = imlec.fetchall()
        imlec.close()
        baglanti.close()
    except Exception:
        return

    client_gemini, koleksiyon = istemcileri_al()

    for kullanici_id, refresh_token, son_gorulen in tum_baglantilar:
        try:
            access_token = await asyncio.to_thread(strava_erisim_tokeni_al, refresh_token)
            aktiviteler = await asyncio.to_thread(strava_son_aktiviteleri_getir, access_token, kac_tane=3)
            for aktivite in reversed(aktiviteler):
                if aktivite["id"] <= (son_gorulen or 0):
                    continue

                aktivite_metni = strava_aktiviteyi_metne_cevir(aktivite)
                yumusak = yumusak_ton_mu(kullanici_id)
                soru = f"Az önce şu antrenmanı bitirdim, yorumlar mısın?\n\n{aktivite_metni}"
                bulunan = await asyncio.to_thread(koleksiyon.query, query_texts=[soru], n_results=KAC_PARCA_GETIRILSIN)
                baglam, _ = baglami_hazirla(bulunan) if bulunan['documents'][0] else ("", [])
                gecmis = gecmisi_oku(kullanici_id)
                tam_profil = tam_profili_olustur(kullanici_id)
                cevap = await asyncio.to_thread(cevap_uret, client_gemini, soru, baglam, gecmis, yumusak=yumusak, profil=tam_profil)

                mesaji_kaydet(kullanici_id, "user", soru)
                mesaji_kaydet(kullanici_id, "model", cevap)
                antrenman_kaydet_tekrarsiz(
                    kullanici_id, f"[Strava-otomatik] {aktivite_metni}",
                    aktivite.get("start_date_local", str(aktivite["id"])),
                )

                await guvenli_send_message(context.bot, kullanici_id, f"🏃 Yeni antrenman algılandı!\n\n{cevap}")
                strava_son_gorulen_guncelle(kullanici_id, aktivite["id"])
        except Exception as e:
            print(f"Strava kontrol hatası (kullanıcı {kullanici_id}): {e}")


async def intervals_kontrol_isi(context: ContextTypes.DEFAULT_TYPE):
    """strava_kontrol_isi ile AYNI mantık — ama Intervals.icu için.
    Huawei'nin native Intervals.icu bağlantısı, ağırlık antrenmanı gibi
    Strava'nın yakalayamadığı türleri de yakalıyor, bu yüzden ayrı
    bir kontrol döngüsü gerekiyor. Her aktivite ID'si AYRI AYRI, kalıcı
    olarak 'bildirildi' diye işaretleniyor — tek bir 'son görülen'
    değişkeni tutmuyoruz, bu eski/yeni aktiviteler arasında sonsuz
    tekrar (ping-pong) yaratıyordu."""
    if not DATABASE_URL:
        return
    try:
        baglanti = psycopg2.connect(DATABASE_URL)
        imlec = baglanti.cursor()
        imlec.execute("SELECT kullanici_id, api_key FROM intervals_baglantilar")
        tum_baglantilar = imlec.fetchall()
        imlec.close()
        baglanti.close()
    except Exception:
        return

    client_gemini, koleksiyon = istemcileri_al()

    for kullanici_id, api_key in tum_baglantilar:
        try:
            aktiviteler = await asyncio.to_thread(intervals_aktiviteleri_getir, api_key, 3)
            for aktivite in reversed(aktiviteler):
                aktivite_id = str(aktivite.get("id", ""))
                if not aktivite_id:
                    continue
                if intervals_aktivite_daha_once_bildirildi_mi(kullanici_id, aktivite_id):
                    continue

                aktivite_metni = intervals_aktiviteyi_metne_cevir(aktivite)
                yumusak = yumusak_ton_mu(kullanici_id)
                soru = f"Az önce şu antrenmanı bitirdim, yorumlar mısın?\n\n{aktivite_metni}"
                bulunan = await asyncio.to_thread(koleksiyon.query, query_texts=[soru], n_results=KAC_PARCA_GETIRILSIN)
                baglam, _ = baglami_hazirla(bulunan) if bulunan['documents'][0] else ("", [])
                gecmis = gecmisi_oku(kullanici_id)
                tam_profil = tam_profili_olustur(kullanici_id)
                cevap = await asyncio.to_thread(cevap_uret, client_gemini, soru, baglam, gecmis, yumusak=yumusak, profil=tam_profil)

                mesaji_kaydet(kullanici_id, "user", soru)
                mesaji_kaydet(kullanici_id, "model", cevap)
                antrenman_kaydet_tekrarsiz(
                    kullanici_id, f"[Intervals.icu-otomatik] {aktivite_metni}",
                    aktivite.get("start_date_local", aktivite_id),
                )

                await guvenli_send_message(context.bot, kullanici_id, f"🏋️ Yeni antrenman algılandı (Intervals.icu)!\n\n{cevap}")
                intervals_aktiviteyi_bildirildi_isaretle(kullanici_id, aktivite_id)
        except Exception as e:
            print(f"Intervals.icu kontrol hatası (kullanıcı {kullanici_id}): {e}")


def _hybrid_arama(koleksiyon, soru, kac_tane):
    """Normal anlamsal aramaya ek olarak, sorudaki net ifadeleri
    (örn. '5. hafta', '3. gün', 'ilk gün', 'üçüncü hafta') KELİME
    OLARAK da arar ve sonuçları birleştirir. Bu, 'Zone 2' ya da
    'X. hafta' gibi çok net ama anlamsal aramanın bazen kaçırdığı
    ifadeleri yakalamayı sağlar."""
    semantik = koleksiyon.query(query_texts=[soru], n_results=kac_tane)
    dokumanlar = list(semantik['documents'][0]) if semantik['documents'][0] else []
    metadatalar = list(semantik['metadatas'][0]) if semantik['metadatas'][0] else []
    gorulen_idler = set(semantik['ids'][0]) if semantik.get('ids') and semantik['ids'][0] else set()

    anahtar_ifadeler = re.findall(r"\d+\s*\.\s*(?:hafta|gün|hafta\w*|gün\w*)", soru, flags=re.IGNORECASE)

    # Yazıyla ifade edilen sıra sayılarını da rakama çevirip aynı şekilde ara
    YAZILI_SAYILAR = {
        "ilk": "1", "birinci": "1", "ikinci": "2", "üçüncü": "3",
        "dördüncü": "4", "beşinci": "5", "altıncı": "6", "yedinci": "7",
        "sekizinci": "8",
    }
    soru_kucuk = soru.lower()
    for kelime, rakam in YAZILI_SAYILAR.items():
        if kelime in soru_kucuk:
            if "gün" in soru_kucuk or "antrenman" in soru_kucuk:
                anahtar_ifadeler.append(f"{rakam}. gün")
            if "hafta" in soru_kucuk:
                anahtar_ifadeler.append(f"{rakam}. hafta")

    for ifade in anahtar_ifadeler[:2]:
        try:
            anahtar_sonuc = koleksiyon.get(
                where_document={"$contains": ifade.strip()}, limit=3,
            )
            for i, doc_id in enumerate(anahtar_sonuc.get("ids", [])):
                if doc_id not in gorulen_idler:
                    gorulen_idler.add(doc_id)
                    dokumanlar.append(anahtar_sonuc["documents"][i])
                    metadatalar.append(anahtar_sonuc["metadatas"][i])
        except Exception:
            continue

    if not dokumanlar:
        return {"documents": [[]], "metadatas": [[]]}
    return {"documents": [dokumanlar], "metadatas": [metadatalar]}


async def _soruyu_isle(update, context, soru, gorsel_b64=None, gorsel_mime=None):
    try:
        client_gemini, koleksiyon = istemcileri_al()
        kullanici_id = update.effective_user.id
        yumusak = yumusak_ton_mu(kullanici_id)

        bulunan = await asyncio.to_thread(_hybrid_arama, koleksiyon, soru, KAC_PARCA_GETIRILSIN)
        baglam, kaynaklar = "", []
        if bulunan['documents'][0]:
            baglam, kaynaklar = baglami_hazirla(bulunan)

        zorla_video_id = context.user_data.get("zorla_video")
        if zorla_video_id:
            try:
                zorla_sonuc = await asyncio.to_thread(koleksiyon.get, where={"video_id": zorla_video_id})
                if zorla_sonuc and zorla_sonuc.get("documents"):
                    zorla_metin = "\n\n".join(zorla_sonuc["documents"])
                    baglam = (
                        f"[ÖNEMLİ — kullanıcı bu videoyu özellikle belirtti, "
                        f"MUTLAKA dikkate al: {zorla_video_id}]\n{zorla_metin}\n\n---\n\n"
                    ) + baglam
            except Exception:
                pass

        # Soru pace/tempo/interval ile ilgiliyse, Strava bağlıysa gerçek pace
        # verilerini OTOMATİK olarak çekip bağlama ekle — kullanıcının ayrıca
        # /strava_ozet çalıştırmasına gerek kalmadan.
        if _kosu_sorusu_mu(soru):
            # Önce Intervals.icu'yu dene — artık antrenmanla ilgili her şeyin
            # (ağırlık + koşu + uyku) tek kaynaktan gelmesi tutarlılık sağlıyor.
            # Intervals.icu bağlı değilse Strava'ya düş.
            intervals_key = intervals_api_key_getir(kullanici_id)
            pace_ozeti = ""
            kaynak_adi = ""
            if intervals_key:
                try:
                    pace_ozeti = await asyncio.to_thread(intervals_kosu_pace_ozeti, intervals_key)
                    kaynak_adi = "Intervals.icu"
                except Exception:
                    pass

            if not pace_ozeti:
                baglanti_bilgisi = strava_baglantisini_getir(kullanici_id)
                if baglanti_bilgisi:
                    try:
                        erisim_tokeni = await asyncio.to_thread(strava_erisim_tokeni_al, baglanti_bilgisi["refresh_token"])
                        pace_ozeti = await asyncio.to_thread(strava_kosu_pace_ozeti, erisim_tokeni)
                        kaynak_adi = "Strava"
                    except Exception:
                        pass

            if pace_ozeti:
                baglam = (
                    f"[GERÇEK {kaynak_adi.upper()} VERİSİ — kullanıcının son koşularının gerçek pace "
                    f"değerleri, öneri verirken buna dayan]\n{pace_ozeti}\n\n---\n\n"
                ) + baglam

        gecmis = gecmisi_oku(kullanici_id)
        profil = tam_profili_olustur(kullanici_id)
        cevap = await asyncio.to_thread(
            cevap_uret, client_gemini, soru, baglam, gecmis, gorsel_b64, gorsel_mime, yumusak, profil
        )

        mesaji_kaydet(kullanici_id, "user", soru)
        mesaji_kaydet(kullanici_id, "model", cevap)

        context.chat_data["son_cevap"] = cevap
        context.chat_data["son_kaynaklar"] = kaynaklar

        dugme_satiri = [
            InlineKeyboardButton("📅 Takvime Hazırla", callback_data="takvim"),
            InlineKeyboardButton("📊 Excel Yap", callback_data="excel"),
        ]
        if kaynaklar:
            dugme_satiri.append(InlineKeyboardButton("🔍 Kaynaklar", callback_data="kaynaklar"))
        dugmeler = InlineKeyboardMarkup([dugme_satiri])
        await guvenli_reply(update.message, cevap, reply_markup=dugmeler)

        # Cevap kullanıcıya gönderildikten SONRA, arka planda (bloklamadan)
        # profili güncelle — kullanıcı bunun bitmesini beklemesin.
        asyncio.create_task(asyncio.to_thread(profili_otomatik_guncelle, client_gemini, kullanici_id, soru, cevap))
        asyncio.create_task(asyncio.to_thread(antrenman_gunlugunu_otomatik_guncelle, client_gemini, kullanici_id, soru, cevap))

        if sesli_cevap_mi(kullanici_id):
            try:
                ses_verisi = await asyncio.to_thread(metni_sese_cevir, client_gemini, cevap)
                if ses_verisi:
                    await update.message.reply_audio(
                        audio=InputFile(BytesIO(ses_verisi), filename="cevap.wav")
                    )
            except Exception as e:
                print(f"Sesli cevap gönderilirken hata: {e}")
    except Exception as e:
        print(f"_soruyu_isle genel hatası: {e}")
        try:
            await update.message.reply_text(
                "Üzgünüm, cevap üretirken beklenmedik bir sorun oluştu. Tekrar dener misin?"
            )
        except Exception:
            pass


async def zorla_video_ayarla(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Kullanım: /zorla_video <video_id_veya_link> — bu videoyu bundan
    sonraki TÜM sorularında garanti olarak dikkate alır (arama şansına
    bırakmaz). Kapatmak için: /zorla_video kapat"""
    if not context.args:
        mevcut = context.user_data.get("zorla_video")
        await update.message.reply_text(
            f"Şu an aktif: {mevcut}" if mevcut else
            "Kullanım: /zorla_video <video_id_veya_link>\nKapatmak için: /zorla_video kapat"
        )
        return

    girdi = context.args[0]
    if girdi.lower() == "kapat":
        context.user_data.pop("zorla_video", None)
        await update.message.reply_text("Zorla video kapatıldı, normal aramaya döndük.")
        return

    video_id = girdi
    eslesme = re.search(r"watch\?v=([\w-]+)", girdi)
    if eslesme:
        video_id = eslesme.group(1)

    context.user_data["zorla_video"] = video_id
    await update.message.reply_text(
        f"Tamamdır, '{video_id}' videosunu bundan sonraki sorularında garanti "
        f"olarak dikkate alacağım. Kapatmak için: /zorla_video kapat"
    )


async def mesaj_geldi(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if "profil_onboarding_index" in context.user_data:
        await _onboarding_cevabini_isle(update, context)
        return
    if "olcum_onboarding_index" in context.user_data:
        await _olcum_cevabini_isle(update, context)
        return
    await _soruyu_isle(update, context, update.message.text)


async def ses_geldi(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        client_gemini, _ = istemcileri_al()
        dosya = await context.bot.get_file(update.message.voice.file_id)
        ses_bytes = bytes(await dosya.download_as_bytearray())
        yazi = await asyncio.to_thread(ses_yaziya_cevir, client_gemini, ses_bytes, "audio/ogg")
        if not yazi:
            await update.message.reply_text("Ses anlaşılamadı, tekrar dener misin?")
            return
        await update.message.reply_text(f"🎤 Anladığım: \"{yazi}\"")
        await _soruyu_isle(update, context, yazi)
    except Exception as e:
        print(f"Sesli mesaj işlenirken hata: {e}")
        await update.message.reply_text(
            "Sesli mesajını işlerken bir sorun oluştu, tekrar dener misin?"
        )


async def foto_geldi(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        dosya = await context.bot.get_file(update.message.photo[-1].file_id)
        foto_bytes = bytes(await dosya.download_as_bytearray())

        if context.user_data.get("yemek_modu"):
            await _yemek_fotografini_isle(update, context, foto_bytes)
            return

        gorsel_b64 = base64.b64encode(foto_bytes).decode("utf-8")
        soru = update.message.caption or "Bu fotoğrafa bakıp yorumlar mısın?"
        await _soruyu_isle(update, context, soru, gorsel_b64, "image/jpeg")
    except Exception as e:
        print(f"Fotoğraf işlenirken hata: {e}")
        await update.message.reply_text(
            "Fotoğrafını işlerken bir sorun oluştu, tekrar dener misin?"
        )


async def buton_tiklandi(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    try:
        await query.answer()
    except Exception as e:
        print(f"query.answer() başarısız (muhtemelen süresi dolmuş), devam ediliyor: {e}")

    client_gemini, _ = istemcileri_al()
    son_cevap = context.chat_data.get("son_cevap", "")
    if not son_cevap:
        await query.message.reply_text("Önce bir program oluşturmam lazım.")
        return

    if query.data == "takvim":
        etkinlikler = await asyncio.to_thread(programdan_json_cikar, client_gemini, son_cevap)
        if not etkinlikler:
            await query.message.reply_text("Bu cevapta takvime çevrilecek bir program bulamadım.")
            return
        ics_veri = ics_plan_olustur(etkinlikler)
        await query.message.reply_document(
            document=InputFile(BytesIO(ics_veri.encode("utf-8")), filename="program.ics"),
            caption="📅 Takvim dosyan hazır, Google Takvim'e aktarabilirsin."
        )
    elif query.data == "excel":
        satirlar = await asyncio.to_thread(programdan_excel_json_cikar, client_gemini, son_cevap)
        if not satirlar:
            await query.message.reply_text("Bu cevapta Excel'e çevrilecek bir program bulamadım.")
            return
        excel_veri = excel_plan_olustur(satirlar)
        await query.message.reply_document(
            document=InputFile(BytesIO(excel_veri), filename="program.xlsx"),
            caption="📊 Excel dosyan hazır."
        )
    elif query.data == "kaynaklar":
        kaynaklar = context.chat_data.get("son_kaynaklar", [])
        if not kaynaklar:
            await query.message.reply_text("Bu cevap için gösterilecek kaynak yok.")
            return
        kaynak_satirlari = []
        gorulenler = set()
        for k in kaynaklar:
            anahtar = k.get("link") or k.get("baslik")
            if not anahtar or anahtar in gorulenler:
                continue
            gorulenler.add(anahtar)
            if k.get("link"):
                kaynak_satirlari.append(f"🎬 {k['link']}")
            elif k.get("baslik"):
                kaynak_satirlari.append(f"📜 {k['baslik']}")
            if len(kaynak_satirlari) >= 8:
                break
        metin = "🔍 Kullanılan kaynaklar (en alakalı olanlar):\n" + "\n".join(kaynak_satirlari)
        await guvenli_reply(query.message, metin)


async def _ders_programini_isle(update: Update, context: ContextTypes.DEFAULT_TYPE, belge):
    """Bir ders/iş programı .xlsx dosyasını okur, her hafta içi günün
    en erken başlangıç / en geç bitiş saatini çıkarır, bunu kalıcı
    profile ekler (böylece TÜM gelecekteki antrenman önerileri bu
    programa göre ayarlanır) ve hemen uygun bir haftalık şablon önerir."""
    import re as _re
    from collections import defaultdict as _defaultdict
    import openpyxl as _openpyxl

    kullanici_id = update.effective_user.id
    dosya = await context.bot.get_file(belge.file_id)
    icerik_bytes = bytes(await dosya.download_as_bytearray())

    try:
        gecici_dosya = BytesIO(icerik_bytes)
        wb = _openpyxl.load_workbook(gecici_dosya, data_only=True)
        ws = wb.active
    except Exception as e:
        await update.message.reply_text(f"Excel dosyası okunamadı: {e}")
        return

    gun_regex = _re.compile(r'^(Monday|Tuesday|Wednesday|Thursday|Friday|Saturday|Sunday|'
                             r'Pazartesi|Salı|Çarşamba|Perşembe|Cuma|Cumartesi|Pazar)\s*[-–]\s*\d')
    zaman_regex = _re.compile(r'^(\d{1,2}):(\d{2})\s*[-–]\s*(\d{1,2}):(\d{2})$')
    gun_ceviri = {
        "Monday": "Pazartesi", "Tuesday": "Salı", "Wednesday": "Çarşamba",
        "Thursday": "Perşembe", "Friday": "Cuma", "Saturday": "Cumartesi", "Sunday": "Pazar",
        "Pazartesi": "Pazartesi", "Salı": "Salı", "Çarşamba": "Çarşamba",
        "Perşembe": "Perşembe", "Cuma": "Cuma", "Cumartesi": "Cumartesi", "Pazar": "Pazar",
    }

    gun_araliklari = _defaultdict(list)
    mevcut_gun = None
    for row in ws.iter_rows(min_row=1, values_only=True):
        if not row or row[0] is None:
            continue
        a_val = str(row[0]).strip()
        gun_match = gun_regex.match(a_val)
        if gun_match:
            mevcut_gun = gun_ceviri.get(gun_match.group(1))
            continue
        zaman_match = zaman_regex.match(a_val)
        if zaman_match and mevcut_gun:
            h1, m1, h2, m2 = map(int, zaman_match.groups())
            gun_araliklari[mevcut_gun].append((h1 * 60 + m1, h2 * 60 + m2))

    if not gun_araliklari:
        await update.message.reply_text(
            "Bu Excel dosyasının yapısını tanıyamadım (gün/saat kalıbı bulunamadı). "
            "Programını kısaca yazarak anlatabilirsin, onu da not alırım."
        )
        return

    gun_sirasi = ["Pazartesi", "Salı", "Çarşamba", "Perşembe", "Cuma", "Cumartesi", "Pazar"]
    ozet_satirlari = ["📅 Ders/iş programımdan çıkarılan haftalık müsaitlik durumum:"]
    for gun in gun_sirasi:
        araliklar = gun_araliklari.get(gun, [])
        if not araliklar:
            ozet_satirlari.append(f"- {gun}: Programda hiç kayıt yok, muhtemelen BOŞ.")
            continue
        en_erken = min(a[0] for a in araliklar)
        en_gec = max(a[1] for a in araliklar)
        ozet_satirlari.append(
            f"- {gun}: Genelde {en_erken // 60:02d}:{en_erken % 60:02d} - "
            f"{en_gec // 60:02d}:{en_gec % 60:02d} arası dolu (ders/klinik)."
        )
    ozet_metni = "\n".join(ozet_satirlari)

    # Kalıcı profile ekle — TÜM gelecekteki antrenman önerileri buna göre ayarlansın
    mevcut_profil = profili_oku(kullanici_id)
    yeni_profil = (mevcut_profil + "\n\n" + ozet_metni).strip() if mevcut_profil else ozet_metni
    profili_yaz(kullanici_id, yeni_profil)

    await update.message.reply_text(
        "✅ Ders programın okundu ve kalıcı profiline eklendi — "
        "bundan sonraki tüm antrenman önerilerimde bunu dikkate alacağım.\n\n" + ozet_metni
    )

    # Hemen uygun bir haftalık antrenman şablonu da önerelim
    try:
        client_gemini, koleksiyon = istemcileri_al()
        yumusak = yumusak_ton_mu(kullanici_id)
        soru = (
            f"Az önce ders/iş programımı ekledim:\n{ozet_metni}\n\n"
            f"Bu programa göre, mevcut antrenman rutinimi (koşu + ağırzık) "
            f"hangi gün/saatlere yerleştirmemi önerirsin? Kısa, pratik bir "
            f"haftalık şablon öner."
        )
        bulunan = await asyncio.to_thread(koleksiyon.query, query_texts=[soru], n_results=KAC_PARCA_GETIRILSIN)
        baglam, _ = baglami_hazirla(bulunan) if bulunan['documents'][0] else ("", [])
        gecmis = gecmisi_oku(kullanici_id)
        cevap = await asyncio.to_thread(
            cevap_uret, client_gemini, soru, baglam, gecmis, yumusak=yumusak, profil=yeni_profil
        )
        mesaji_kaydet(kullanici_id, "user", soru)
        mesaji_kaydet(kullanici_id, "model", cevap)
        await guvenli_reply(update.message, cevap)
    except Exception as e:
        print(f"Ders programına göre şablon önerilirken hata: {e}")


async def _md_dosyasini_isle(update: Update, context: ContextTypes.DEFAULT_TYPE, belge):
    """Bir video transkripti (.md) dosyasını doğrudan canlı ChromaDB
    arşivine ekler — zip yükleme/redeploy derdi olmadan, anında."""
    dosya = await context.bot.get_file(belge.file_id)
    icerik_bytes = bytes(await dosya.download_as_bytearray())
    try:
        metin = icerik_bytes.decode("utf-8")
    except Exception as e:
        await update.message.reply_text(f"Dosya okunamadı: {e}")
        return

    if len(metin.strip()) < 50:
        await update.message.reply_text("Dosya çok kısa/boş görünüyor, atlandı.")
        return

    # "# Video Linki: https://youtube.com/watch?v=XXXX" satırından video_id çıkar
    video_id = belge.file_name.replace(".md", "")
    eslesme = re.search(r"watch\?v=([\w-]+)", metin)
    if eslesme:
        video_id = eslesme.group(1)

    try:
        _, koleksiyon = istemcileri_al()

        boyut, ortusme = 800, 150
        parcalar, baslangic = [], 0
        while baslangic < len(metin):
            parcalar.append(metin[baslangic:baslangic + boyut])
            baslangic += (boyut - ortusme)

        # Bu video daha önce eklenmişse eski kayıtlarını temizle (tekrar etmesin)
        try:
            eski = koleksiyon.get(where={"video_id": video_id}, include=[])
            if eski.get("ids"):
                koleksiyon.delete(ids=eski["ids"])
        except Exception:
            pass

        ids = [f"{video_id}_parca_{j}" for j in range(len(parcalar))]
        metadatalar = [{"video_id": video_id, "kaynak": belge.file_name} for _ in parcalar]
        koleksiyon.add(documents=parcalar, ids=ids, metadatas=metadatalar)

        await update.message.reply_text(
            f"✅ '{video_id}' videosu canlı arşive eklendi ({len(parcalar)} parça). "
            f"Artık bu içerikten sorular sorabilirsin."
        )
    except Exception as e:
        await update.message.reply_text(f"Arşive eklenirken hata oluştu: {e}")


async def belge_geldi(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """.json (eski sohbet), .md (video transkripti) ya da .xlsx (ders
    programı) dosyası gönderildiğinde işler."""
    belge = update.message.document

    if belge.file_name.endswith(".md"):
        await _md_dosyasini_isle(update, context, belge)
        return

    if belge.file_name.endswith(".xlsx"):
        await _ders_programini_isle(update, context, belge)
        return

    if not belge.file_name.endswith(".json"):
        await update.message.reply_text(
            "Şu an .json (eski sohbet), .md (video transkripti) ya da "
            ".xlsx (ders programı) dosyası kabul ediyorum."
        )
        return

    dosya = await context.bot.get_file(belge.file_id)
    icerik_bytes = bytes(await dosya.download_as_bytearray())

    try:
        veri = json.loads(icerik_bytes.decode("utf-8"))
    except Exception as e:
        await update.message.reply_text(f"Dosya okunamadı: {e}")
        return

    mesajlar = veri.get("mesajlar", [])
    if not mesajlar:
        await update.message.reply_text("Bu dosyada mesaj bulunamadı, atlandı.")
        return

    baslik = veri.get("baslik", belge.file_name)
    kullanici_id = update.effective_user.id

    # 1) Arşive (ChromaDB) ekle — bilgi olarak her zaman aranabilir olsun
    try:
        _, koleksiyon = istemcileri_al()
        satirlar = [f"# Eski Sohbet: {baslik}\n"]
        for m in mesajlar:
            rol = "Kullanıcı" if m.get("role") == "user" else "Koçum"
            icerik = m.get("content", "")
            if icerik:
                satirlar.append(f"{rol}: {icerik}")
        metin = "\n\n".join(satirlar)

        parca_sayisi = 0
        if len(metin.strip()) >= 50:
            boyut, ortusme = 800, 150
            parcalar, baslangic = [], 0
            while baslangic < len(metin):
                parcalar.append(metin[baslangic:baslangic + boyut])
                baslangic += (boyut - ortusme)
            etiket = f"eski_sohbet_{uuid.uuid4().hex[:8]}"
            ids = [f"{etiket}_parca_{j}" for j in range(len(parcalar))]
            metadatalar = [{"video_id": f"Eski sohbet: {baslik}", "kaynak": "eski_sohbet"}
                            for _ in parcalar]
            koleksiyon.add(documents=parcalar, ids=ids, metadatas=metadatalar)
            parca_sayisi = len(parcalar)
    except Exception as e:
        parca_sayisi = f"HATA: {e}"

    # 2) GERÇEK konuşma geçmişine ekle — böylece bota "geçen sohbette
    #    ne demiştik" dediğinde arama yapmadan doğrudan hatırlar
    eklenen_mesaj = 0
    for m in mesajlar:
        icerik = m.get("content", "")
        rol = m.get("role")
        if not icerik:
            continue
        mesaji_kaydet(kullanici_id, "user" if rol == "user" else "model", icerik)
        eklenen_mesaj += 1

    await update.message.reply_text(
        f"📥 '{baslik}' eklendi:\n"
        f"- Arşive: {parca_sayisi} parça\n"
        f"- Gerçek sohbet geçmişine: {eklenen_mesaj} mesaj — artık doğrudan hatırlıyorum, "
        f"arama yapmama bile gerek yok."
    )


def main():
    if not TELEGRAM_BOT_TOKEN:
        print("HATA: TELEGRAM_BOT_TOKEN ayarlanmamış.")
        return
    _veritabanini_hazirla()
    _basit_semayi_hazirla()
    istemcileri_al()

    app = Application.builder().token(TELEGRAM_BOT_TOKEN).build()
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("id", id_goster))
    app.add_handler(CommandHandler("model_listesi", model_listesi))
    app.add_handler(CommandHandler("yumusak_ac", yumusak_ac))
    app.add_handler(CommandHandler("yumusak_kapat", yumusak_kapat))
    app.add_handler(CommandHandler("temizle", temizle))
    app.add_handler(CommandHandler("web_sohbetlerini_getir", web_sohbetlerini_getir))
    app.add_handler(CommandHandler("strava_baglan", strava_baglan))
    app.add_handler(CommandHandler("son_antrenman", son_antrenman))
    app.add_handler(CommandHandler("video_var_mi", video_var_mi))
    app.add_handler(CommandHandler("zorla_video", zorla_video_ayarla))
    app.add_handler(CommandHandler("profil_goster", profil_goster))
    app.add_handler(CommandHandler("profil_ekle", profil_ekle))
    app.add_handler(CommandHandler("profil_sil", profil_sil))
    app.add_handler(CommandHandler("profil_olustur", profil_olustur))
    app.add_handler(CommandHandler("iptal", profil_iptal))
    app.add_handler(CommandHandler("strava_ozet", strava_ozet))
    app.add_handler(CommandHandler("sabah_ac", sabah_ac))
    app.add_handler(CommandHandler("sabah_kapat", sabah_kapat))
    app.add_handler(CommandHandler("olcum_ekle", olcum_ekle))
    app.add_handler(CommandHandler("olcum_gecmisi", olcum_gecmisi))
    app.add_handler(CommandHandler("olcum_hatirlatma_ac", olcum_hatirlatma_ac))
    app.add_handler(CommandHandler("olcum_hatirlatma_kapat", olcum_hatirlatma_kapat))
    app.add_handler(CommandHandler("sesli_cevap_ac", sesli_cevap_ac))
    app.add_handler(CommandHandler("sesli_cevap_kapat", sesli_cevap_kapat))
    app.add_handler(CommandHandler("yemek_ekle", yemek_ekle))
    app.add_handler(CommandHandler("intervals_baglan", intervals_baglan))
    app.add_handler(CommandHandler("uyku_durumu", uyku_durumu))
    app.add_handler(CommandHandler("antrenman_gecmisi", antrenman_gecmisi))
    app.add_handler(CommandHandler("antrenman_ekle", antrenman_ekle))
    app.add_handler(CommandHandler("beslenme_ozet", beslenme_ozet))
    app.add_handler(CommandHandler("yemek_duzelt", yemek_duzelt))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, mesaj_geldi))
    app.add_handler(MessageHandler(filters.VOICE, ses_geldi))
    app.add_handler(MessageHandler(filters.PHOTO, foto_geldi))
    app.add_handler(MessageHandler(filters.Document.ALL, belge_geldi))
    app.add_handler(CallbackQueryHandler(buton_tiklandi))

    # Her 15 dakikada bir, tüm bağlı kullanıcıların yeni Strava aktivitesi
    # olup olmadığını kontrol eder — her kullanıcıya SADECE KENDİ verisi gider.
    if app.job_queue:
        # NOT: strava_kontrol_isi (otomatik Strava bildirimi) BİLEREK
        # kapatıldı — Huawei artık aynı antrenmanı hem Strava'ya hem
        # Intervals.icu'ya gönderdiği için, ikisini de otomatik izlemek
        # ÇİFT bildirime sebep oluyordu. Intervals.icu artık her şeyi
        # (koşu/yürüyüş + ağırlık antrenmanı) tek kaynaktan yakalıyor.
        # Strava hâlâ /strava_ozet ve pace algılama için kullanılıyor.
        # app.job_queue.run_repeating(strava_kontrol_isi, interval=900, first=30)
        app.job_queue.run_repeating(intervals_kontrol_isi, interval=900, first=45)

        # Sabit 07:00 yerine, her 30 dakikada bir kontrol eder — o günün
        # uyku verisi gerçekten hazır olunca (kullanıcı muhtemelen kalkmış
        # demektir) mesajı gönderir, günde sadece 1 kez. Veri yoksa ve
        # saat henüz 11:00'i geçmediyse bekler (fonksiyonun kendi içinde
        # kontrol ediliyor).
        try:
            from zoneinfo import ZoneInfo
            from datetime import time as _time
            app.job_queue.run_repeating(sabah_mesaji_isi, interval=1800, first=60)
            # Her gün 09:00'da, son ölçümünden 30+ gün geçmiş kullanıcılara
            # aylık ölçüm hatırlatması gönderir.
            app.job_queue.run_daily(
                olcum_hatirlatma_isi, time=_time(9, 0, tzinfo=ZoneInfo("Europe/Istanbul")),
            )
        except Exception as e:
            print(f"Zamanlanmış işler kurulamadı: {e}")

    print(f"{UYGULAMA_ADI} Telegram botu başlıyor...")
    app.run_polling()


if __name__ == "__main__":
    main()
