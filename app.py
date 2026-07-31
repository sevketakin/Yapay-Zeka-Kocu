"""
Koçum - Chainlit Tabanlı Kişisel Antrenör Asistanı
================================================================

Streamlit'ten Chainlit'e geçiş. Chainlit, LLM sohbet uygulamaları için
özel tasarlanmış bir framework olduğu için, ChatGPT'ye çok benzer,
kutudan çıktığı gibi cilalı bir arayüz, çoklu sohbet geçmişi ve daha
stabil dosya/görsel yükleme sağlıyor.

KURULUM:
    pip install chainlit sqlalchemy aiosqlite chromadb google-genai openpyxl

BİR KEZ YAPILACAK - Kimlik doğrulama anahtarı oluştur:
    chainlit create-secret
    (çıkan satırı .env dosyasına CHAINLIT_AUTH_SECRET=... şeklinde kaydet)

ÇALIŞTIRMA:
    chainlit run app.py -w

Bu komut tarayıcında otomatik olarak bir sekme açacak (localhost:8000).
"""

import os
import re
import json
import uuid
import base64
import zipfile
from datetime import datetime, timedelta
from io import BytesIO

import chainlit as cl
from chainlit.data.sql_alchemy import SQLAlchemyDataLayer
import chromadb
from chromadb import Documents, EmbeddingFunction, Embeddings
from google import genai

# ============== AYARLAR ==============
# VERI_KLASORU: Railway'de kalıcı Volume'un bağlandığı yer (örn. /data).
# Yerelde test ederken normal klasörleri kullanır (Volume yoksa).
VERI_KLASORU = os.environ.get("VERI_KLASORU", ".")
DB_KLASORU = os.path.join(VERI_KLASORU, "veritabani")
SOHBET_DB_YOLU = os.path.join(VERI_KLASORU, "koc_data.db")

KOLEKSIYON_ADI = "video_transkriptleri"
GEMINI_API_ANAHTARI = os.environ.get("GEMINI_API_ANAHTARI", "BURAYA_API_ANAHTARINI_YAPISTIR")
GEMINI_MODEL_LISTESI = [
    "gemini-flash-latest",
    "gemini-3.5-flash",
    "gemini-3.1-flash-lite",
    "gemini-3-flash-preview",
]
KAC_PARCA_GETIRILSIN = 15
UYGULAMA_ADI = "Koçum"
# =======================================

# ============== VERİTABANI ZIP'İNİ OTOMATİK İNDİRME + BİRLEŞTİRME + AÇMA ==============
# Öncelik sırası:
# 1) '/data/veritabani' zaten varsa hiçbir şey yapma
# 2) Yoksa, önce mevcut zip'in SAĞLAM olup olmadığını kontrol et — bozuksa sil
# 3) GitHub Release'ten indir (GITHUB_ZIP_URL tanımlıysa)
# 4) Ya da yerelde yüklenmiş parçaları (veritabani.zip.001, ...) birleştir
# 5) Zip'i aç, açma da bozuksa tekrar sil ve baştan indir (bir kere daha dene)
GITHUB_ZIP_URL = (
    "https://github.com/sevketakin/Yapay-Zeka-Kocu/releases/download/"
    "v1/veritabani.zip"
)


def _zip_saglam_mi(yol):
    try:
        with zipfile.ZipFile(yol, 'r') as z:
            return z.testzip() is None
    except Exception:
        return False


def _veritabanini_hazirla():
    import glob as _glob
    import requests

    _tek_zip = os.path.join(VERI_KLASORU, "veritabani.zip")
    _birlesik_zip = os.path.join(VERI_KLASORU, "_veritabani_birlesik.zip")

    # Var olan ama BOZUK dosyaları temizle (önceki başarısız denemelerden kalmış olabilir)
    for _yol in [_tek_zip, _birlesik_zip]:
        if os.path.exists(_yol) and not _zip_saglam_mi(_yol):
            print(f"'{_yol}' bozuk görünüyor, siliniyor...")
            os.remove(_yol)

    # Eski/yarım kalmış parça dosyalarını da temizle (yer açmak için)
    for _parca in _glob.glob(os.path.join(VERI_KLASORU, "veritabani.zip.*")):
        if not os.path.exists(_tek_zip):  # tek zip zaten sağlamsa parçalara dokunma
            try:
                os.remove(_parca)
            except Exception:
                pass

    if GITHUB_ZIP_URL and not os.path.exists(_tek_zip):
        try:
            print("Veritabanı GitHub Release'ten indiriliyor... (bu birkaç dakika sürebilir)")
            with requests.get(GITHUB_ZIP_URL, stream=True, allow_redirects=True, timeout=120) as yanit:
                yanit.raise_for_status()
                toplam = int(yanit.headers.get("content-length", 0))
                indirilen = 0
                with open(_tek_zip, "wb") as f:
                    for parca in yanit.iter_content(chunk_size=8 * 1024 * 1024):
                        f.write(parca)
                        indirilen += len(parca)
                        if toplam:
                            print(f"  {indirilen / (1024*1024):.0f}MB / {toplam / (1024*1024):.0f}MB")
            print("İndirme tamamlandı.")
        except Exception as e:
            print(f"GitHub'dan indirme başarısız: {e}")
            if os.path.exists(_tek_zip):
                try:
                    os.remove(_tek_zip)
                    print("Yarım kalan dosya temizlendi.")
                except Exception:
                    pass

    if not os.path.exists(_tek_zip) or not _zip_saglam_mi(_tek_zip):
        _parcalar = sorted(_glob.glob(os.path.join(VERI_KLASORU, "veritabani.zip.*")))
        if _parcalar and not os.path.exists(_birlesik_zip):
            print(f"{len(_parcalar)} parça bulundu, birleştiriliyor...")
            with open(_birlesik_zip, "wb") as cikti:
                for parca in _parcalar:
                    with open(parca, "rb") as p:
                        cikti.write(p.read())
            print("Parçalar birleştirildi.")

    _acilacak_zip = _tek_zip if _zip_saglam_mi(_tek_zip) else (
        _birlesik_zip if os.path.exists(_birlesik_zip) and _zip_saglam_mi(_birlesik_zip) else None
    )

    if not _acilacak_zip:
        print("HATA: Sağlam bir veritabanı zip'i bulunamadı/indirilemedi.")
        return

    print(f"'{_acilacak_zip}' açılıyor... (bu biraz sürebilir)")
    with zipfile.ZipFile(_acilacak_zip, 'r') as z:
        z.extractall(VERI_KLASORU)

    # Açma başarılıysa, yer açmak için zip'i (ve varsa parçaları) sil
    try:
        for _yol in [_tek_zip, _birlesik_zip]:
            if os.path.exists(_yol):
                os.remove(_yol)
        for _parca in _glob.glob(os.path.join(VERI_KLASORU, "veritabani.zip.*")):
            os.remove(_parca)
    except Exception:
        pass

    # "Tamamlandı" işareti — bu dosya varsa, açma işleminin SONUNA kadar
    # sorunsuz geldiğini biliyoruz. Yoksa (yarım kalmışsa) bir dahaki
    # başlangıçta her şeyi silip baştan deneriz.
    with open(_tamamlandi_isareti, "w") as f:
        f.write("ok")
    print("Veritabanı hazır.")


_tamamlandi_isareti = os.path.join(VERI_KLASORU, "_veritabani_tamamlandi.txt")

if not os.path.exists(_tamamlandi_isareti):
    # Önceki deneme yarım kalmış olabilir (klasör kısmen oluşmuş ama
    # işaret dosyası yok) — güvenli olması için varsa temizleyip
    # sıfırdan başlıyoruz.
    if os.path.exists(DB_KLASORU):
        print("Yarım kalmış eski veritabanı klasörü bulundu, temizleniyor...")
        import shutil as _shutil
        _shutil.rmtree(DB_KLASORU, ignore_errors=True)
    _veritabanini_hazirla()

AYLAR_TR = ["Ocak", "Şubat", "Mart", "Nisan", "Mayıs", "Haziran",
            "Temmuz", "Ağustos", "Eylül", "Ekim", "Kasım", "Aralık"]
GUNLER_TR = ["Pazartesi", "Salı", "Çarşamba", "Perşembe", "Cuma", "Cumartesi", "Pazar"]
GUN_INDEX = {"Pazartesi": 0, "Salı": 1, "Çarşamba": 2, "Perşembe": 3,
             "Cuma": 4, "Cumartesi": 5, "Pazar": 6}


class GeminiEmbeddingFonksiyonu(EmbeddingFunction):
    """Veritabanı 'gemini-embedding-001' ile indexlendiği için, arama
    yaparken de AYNI modeli kullanmamız gerekiyor — aksi halde
    embedding'ler uyumsuz olur ve arama anlamsızlaşır."""

    def __init__(self, api_anahtari):
        self.client = genai.Client(api_key=api_anahtari)

    def __call__(self, input: Documents) -> Embeddings:
        return [self._tek_embedding_al(metin) for metin in input]

    def _tek_embedding_al(self, metin, max_deneme=3):
        import time
        for deneme in range(1, max_deneme + 1):
            try:
                sonuc = self.client.models.embed_content(
                    model="gemini-embedding-001",
                    contents=metin,
                )
                return sonuc.embeddings[0].values
            except Exception as e:
                hata_metni = str(e)
                if "429" in hata_metni or "RESOURCE_EXHAUSTED" in hata_metni or "UNAVAILABLE" in hata_metni:
                    time.sleep(5)
                    continue
                raise
        raise RuntimeError(f"Embedding alınamadı: {metin[:50]}...")


def _turkiye_simdi():
    try:
        from zoneinfo import ZoneInfo
        return datetime.now(ZoneInfo("Europe/Istanbul"))
    except Exception:
        return datetime.now()


def bugunun_tarihi():
    simdi = _turkiye_simdi()
    return f"{simdi.day} {AYLAR_TR[simdi.month - 1]} {simdi.year}, {GUNLER_TR[simdi.weekday()]}"


def su_anki_saat():
    simdi = _turkiye_simdi()
    saat = simdi.hour
    if 5 <= saat < 12:
        vakit = "sabah"
    elif 12 <= saat < 17:
        vakit = "öğleden sonra"
    elif 17 <= saat < 21:
        vakit = "akşam"
    else:
        vakit = "gece"
    return simdi.strftime("%H:%M"), vakit


# ============== TEKİL İSTEMCİLER (bir kere oluşturulur) ==============
_client_gemini = None
_koleksiyon = None


def istemcileri_al():
    global _client_gemini, _koleksiyon
    if _client_gemini is None:
        _client_gemini = genai.Client(api_key=GEMINI_API_ANAHTARI)
    if _koleksiyon is None:
        embed_fonksiyonu = GeminiEmbeddingFonksiyonu(GEMINI_API_ANAHTARI)
        db = chromadb.PersistentClient(path=DB_KLASORU)
        _koleksiyon = db.get_collection(name=KOLEKSIYON_ADI, embedding_function=embed_fonksiyonu)
    return _client_gemini, _koleksiyon


def baglami_hazirla(bulunan_parcalar):
    metinler = bulunan_parcalar['documents'][0]
    metadatalar = bulunan_parcalar['metadatas'][0]
    baglam_parcalari = []
    kaynaklar = []
    for metin, meta in zip(metinler, metadatalar):
        video_id = meta.get('video_id', 'bilinmiyor')
        kaynak_turu = meta.get('kaynak', '')
        onizleme = metin[:250].strip() + ("..." if len(metin) > 250 else "")

        if kaynak_turu == "eski_sohbet":
            # Bu GERÇEK bir kişisel konuşma geçmişi — genel bir video değil
            baglam_parcalari.append(
                f"[GERÇEK KİŞİSEL GEÇMİŞ — {video_id} — bu kullanıcının SENİNLE "
                f"gerçekten daha önce konuştuğu bir şey, genel bir video örneği DEĞİL]\n{metin}"
            )
            kaynaklar.append({"link": None, "baslik": video_id, "onizleme": onizleme})
        else:
            link = f"https://youtube.com/watch?v={video_id}"
            baglam_parcalari.append(
                f"[Genel video içeriği (BAŞKA insanların/koçların anlattığı örnekler, "
                f"kullanıcının kendi kişisel geçmişi DEĞİL): {link}]\n{metin}"
            )
            kaynaklar.append({"link": link, "baslik": None, "onizleme": onizleme})
    return "\n\n---\n\n".join(baglam_parcalari), kaynaklar


def ses_yaziya_cevir(client_gemini, ses_bytes, ses_mime_tipi="audio/wav"):
    ses_b64 = base64.b64encode(ses_bytes).decode("utf-8")
    for model_adi in GEMINI_MODEL_LISTESI:
        try:
            yanit = client_gemini.models.generate_content(
                model=model_adi,
                contents=[
                    {"role": "user", "parts": [
                        {"text": "Bu ses kaydını sadece yazıya çevir, başka hiçbir şey ekleme, sadece söylenen metni döndür."},
                        {"inline_data": {"mime_type": ses_mime_tipi, "data": ses_b64}},
                    ]}
                ],
            )
            return yanit.text.strip()
        except Exception:
            continue
    return None


def cevap_uret(client_gemini, soru, baglam, gecmis, gorsel_b64=None, gorsel_mime=None):
    bugun = bugunun_tarihi()
    saat, vakit = su_anki_saat()
    sistem_mesaji = (
        "🚨 EN ÖNEMLİ KURAL — EN BAŞTA OKU VE HER ZAMAN UYGULA:\n"
        "Kullanıcı sana 'geçen hafta/gün ne yaptık', 'dün ne konuşmuştuk', "
        "'hatırlıyor musun', kaldırdığı ağırlıklar, koştuğu mesafeler gibi "
        "KENDİ KİŞİSEL geçmişiyle ilgili bir şey sorduğunda: SADECE ve "
        "SADECE aşağıda sana verilen gerçek konuşma geçmişinde (ÖNCEKİ "
        "MESAJLAR) ya da '[GERÇEK KİŞİSEL GEÇMİŞ ...]' etiketli notlarda "
        "GERÇEKTEN yazan bilgiyi kullan. '[Genel video içeriği ...]' "
        "etiketli notlar BAŞKA insanların (YouTuber'ların, koçların) "
        "kendi hikayeleri — bunlardaki 'geçen hafta yarışa katıldım', "
        "'Kapadokya'ya gittim' gibi ifadeler O VİDEODAKİ KİŞİYE aittir, "
        "SANA SORU SORAN KULLANICIYA DEĞİL. Bunları asla kullanıcının "
        "kendi hikayesiymiş gibi anlatma.\n\n"
        "ÖRNEK — YANLIŞ DAVRANIŞ (ASLA BÖYLE YAPMA):\n"
        "Kullanıcı: 'Geçen hafta ne yaptık?'\n"
        "Sen (YANLIŞ): 'Geçen hafta Kapadokya'ya gittik, yarışta 2. oldun, "
        "Sergio ile yarıştın...' (Bu tamamen uydurma — bir video "
        "transkriptinden alınıp kullanıcıya mal edilmiş, GERÇEK DEĞİL)\n\n"
        "ÖRNEK — DOĞRU DAVRANIŞ:\n"
        "Kullanıcı: 'Geçen hafta ne yaptık?'\n"
        "Sen (DOĞRU, eğer konuşma geçmişinde/gerçek notlarda bilgi yoksa): "
        "'Geçen hafta neler yaptığını bana anlatmamıştın aslında, hatırlatır "
        "mısın? Ona göre bugünkü antrenmanı planlayalım.'\n\n"
        "Bu kural, diğer TÜM talimatlardan daha önceliklidir. Kişisel "
        "veri uydurmak, gerçek bir antrenörün asla yapmayacağı, güven "
        "kırıcı bir hatadır.\n\n"
        f"Bugünün tarihi: {bugun}. Şu anki saat: {saat} ({vakit}). "
        f"Sorulara göre (örneğin bu haftaki antrenman planı, kaç gün kaldı, "
        f"bugün ne yesem, şu an antrenman yapmalı mıyım gibi) bu tarih ve "
        f"saat bilgisini mutlaka dikkate al ve cevabını buna göre uyarla. "
        f"Örneğin: gece geç saatte ağır bir antrenman veya kafeinli bir "
        f"şey önerme, sabah/akşam selamlaşmanı zaman dilimine göre yap, "
        f"'bu akşam' veya 'yarın sabah' gibi ifadeleri gerçek zamana göre "
        f"doğru kullan.\n\n"
        "Sen benim kişisel hybrid antrenörümsün. Amacın, beni hybrid "
        "sistemde (koşu, bisiklet, yüzme, kuvvet vb. dallarda) en iyi "
        "versiyonuma ulaştırmak için elinden gelen tüm yardımı, bilgiyi "
        "ve desteği sağlamak. Spor bilimi, antrenman periyotlaması, "
        "beslenme, toparlanma ve hybrid atletizm konularında yıllarını "
        "vermiş, sahada bizzat çalışmış, deneyimli bir antrenör/sporcu "
        "gibisin.\n\n"
        "Sana ayrıca, senin yıllar içinde izlediğin, öğrendiğin ve "
        "tecrübe edindiğin antrenman videolarından/içeriklerden bazı "
        "hatırlatma notları verilecek. BUNLAR SENİN KENDİ BİLGİ "
        "BİRİKİMİN — dışarıdan bir kaynak, döküman veya alıntı DEĞİL. "
        "Bu bilgiyi asla 'bu videoda deniyor ki', 'transkriptte geçiyor "
        "ki', 'kaynağa göre' gibi ifadelerle aktarma ve asla dışarıdan "
        "bir referans gibi sunma. Bunun yerine, sanki yıllar içinde "
        "bizzat öğrenip özümsediğin, kendi tecrüben ve bilgin haline "
        "gelmiş bir şeymiş gibi, doğal ve kendinden emin bir şekilde "
        "anlat — tıpkı gerçek bir antrenörün, okuduğu/izlediği onlarca "
        "kaynaktan süzdüğü bilgiyi kendi sözleriyle, kendi tecrübesi "
        "gibi aktarması gibi.\n\n"
        "ÖNCELİK SIRASI (ÇOK ÖNEMLİ): Bu notlar soruyla alakalıysa, "
        "cevabını ÖNCELİKLE ve AĞIRLIKLI OLARAK bu notlara dayandır — "
        "onlar senin birincil, en somut ve en güncel bilgi kaynağın gibi "
        "davran. Kendi genel bilgin sadece bu notları TAMAMLAMAK, "
        "boşlukları doldurmak veya bağlam eklemek için ikincil bir "
        "destek olarak kullanılır. Yani alakalı not varken genel "
        "bilgini öne çıkarıp notu arka planda bırakma — tam tersi "
        "olsun. Notlar arasında çelişki varsa ya da soru notlarla hiç "
        "alakalı değilse, o zaman kendi genel uzmanlığından tam ve "
        "kaliteli bir cevap ver.\n\n"
        "🚨 KRİTİK GÜVENLİK KURALI — KULLANICININ KİŞİSEL GEÇMİŞİNİ ASLA "
        "UYDURMA: Yukarıdaki 'kendinden emin konuş' talimatı SADECE genel "
        "spor bilimi/antrenman bilgisi için geçerli — kullanıcının KENDİ "
        "KİŞİSEL antrenman geçmişi, yaptığı spesifik idmanlar, kaldırdığı "
        "ağırlıklar, 'geçen gün ne yaptık', 'dün ne konuşmuştuk' gibi "
        "sorularda KESİNLİKLE FARKLI davran: Bu tür kişisel/geçmişe "
        "dayalı iddiaları SADECE aşağıda sana verilen gerçek konuşma "
        "geçmişinde (önceki mesajlarda) ya da 'Eski sohbet: ...' etiketli "
        "notlarda GERÇEKTEN yazıyorsa kullan. Video transkriptlerindeki "
        "genel örnekleri (başka insanların/koçların anlattığı antrenmanları) "
        "ASLA kullanıcının KENDİ geçmişiymiş gibi sunma — bu, sayı, "
        "ağırlık, tarih gibi somut kişisel detaylar UYDURMAK demektir ve "
        "kesinlikle yasak, çok tehlikeli bir güven kaybına yol açar. "
        "Kullanıcının geçmişte ne yaptığını GERÇEKTEN bilmiyorsan, bunu "
        "olduğu gibi, dürüstçe ve doğal bir dille söyle (örn. 'bunu "
        "bana daha önce anlatmamıştın, hatırlatır mısın?' gibi) — asla "
        "sayı/ağırlık/tarih uydurup kesin bir hatıraymış gibi anlatma.\n\n"
        "TON: Benimle samimi, sıcak, motive edici ve gerçek bir antrenör "
        "gibi konuş — resmi/mesafeli bir asistan gibi değil. Beni "
        "tanıyan, hedeflerime yatırım yapmış biri gibi davran. Önceki "
        "mesajları dikkate alarak sohbetin bağlamını koru."
    )

    contents = list(gecmis)
    kullanici_mesaji = f"""(Senin kendi bilgi birikimin/hafızandan gelen, bu soruyla \
alakalı olabilecek notlar — alakasızsa yok say, alakalıysa kendi tecrübenmiş gibi kullan):

{baglam}

SORU: {soru}"""

    parts = [{"text": kullanici_mesaji}]
    if gorsel_b64:
        parts.append({"inline_data": {"mime_type": gorsel_mime, "data": gorsel_b64}})
    contents.append({"role": "user", "parts": parts})

    son_hata = None
    for model_adi in GEMINI_MODEL_LISTESI:
        for deneme in range(1, 3):
            try:
                yanit = client_gemini.models.generate_content(
                    model=model_adi,
                    contents=contents,
                    config={"system_instruction": sistem_mesaji},
                )
                metin = yanit.text
                if metin and metin.strip():
                    return metin
                son_hata = "Model boş cevap döndürdü"
                break
            except Exception as e:
                hata_metni = str(e)
                son_hata = hata_metni
                if "503" in hata_metni or "UNAVAILABLE" in hata_metni:
                    continue
                elif "404" in hata_metni or "NOT_FOUND" in hata_metni:
                    break
                else:
                    continue

    return (f"Şu an cevap üretemedim, lütfen tekrar dener misin? "
            f"(Teknik detay: {son_hata})")


# ============== PROGRAM -> TAKVİM (.ics) DÖNÜŞÜMÜ ==============
def programdan_json_cikar(client_gemini, program_metni):
    talimat = (
        "Aşağıdaki metinde bir antrenman/beslenme/aktivite programı var. "
        "Bunu SADECE aşağıdaki JSON formatında, başka hiçbir açıklama/metin "
        "eklemeden döndür:\n\n"
        '[{"gun": "Pazartesi", "baslik": "Bacak Antrenmanı", '
        '"baslangic_saat": "18:00", "bitis_saat": "19:00", '
        '"aciklama": "kısa açıklama"}, ...]\n\n'
        "Kurallar:\n"
        "- gun değeri sadece şunlardan biri olmalı: Pazartesi, Salı, Çarşamba, "
        "Perşembe, Cuma, Cumartesi, Pazar\n"
        "- Saat belirtilmemişse mantıklı bir saat öner\n"
        "- Metinde gerçekten gün/aktivite içeren bir program yoksa boş liste "
        "[] döndür\n\n"
        f"METİN:\n{program_metni}"
    )
    for model_adi in GEMINI_MODEL_LISTESI:
        try:
            yanit = client_gemini.models.generate_content(model=model_adi, contents=talimat)
            metin = yanit.text.strip()
            metin = re.sub(r"^```json\s*|\s*```$", "", metin.strip(), flags=re.MULTILINE)
            metin = metin.strip("`").strip()
            return json.loads(metin)
        except Exception:
            continue
    return None


def sonraki_gun_tarihi(hedef_gun_index):
    bugun = datetime.now().date()
    fark = (hedef_gun_index - bugun.weekday()) % 7
    return bugun + timedelta(days=fark)


def ics_plan_olustur(etkinlikler, baslik_oneki=""):
    satirlar = ["BEGIN:VCALENDAR", "VERSION:2.0", "PRODID:-//Kocum//TR//", "CALSCALE:GREGORIAN"]
    for e in etkinlikler:
        gun = e.get("gun", "")
        if gun not in GUN_INDEX:
            continue
        try:
            bs_saat, bs_dk = map(int, e.get("baslangic_saat", "18:00").split(":"))
            bt_saat, bt_dk = map(int, e.get("bitis_saat", "19:00").split(":"))
        except Exception:
            continue
        tarih = sonraki_gun_tarihi(GUN_INDEX[gun])
        uid = str(uuid.uuid4())
        baslik = e.get("baslik", "Etkinlik")
        if baslik_oneki:
            baslik = f"[{baslik_oneki}] {baslik}"
        satirlar.append("BEGIN:VEVENT")
        satirlar.append(f"UID:{uid}")
        satirlar.append(f"DTSTAMP:{datetime.now().strftime('%Y%m%dT%H%M%SZ')}")
        satirlar.append(f"DTSTART;TZID=Europe/Istanbul:{tarih.strftime('%Y%m%d')}T{bs_saat:02d}{bs_dk:02d}00")
        satirlar.append(f"DTEND;TZID=Europe/Istanbul:{tarih.strftime('%Y%m%d')}T{bt_saat:02d}{bt_dk:02d}00")
        satirlar.append(f"SUMMARY:{baslik}")
        if e.get("aciklama"):
            satirlar.append(f"DESCRIPTION:{e['aciklama']}")
        satirlar.append("END:VEVENT")
    satirlar.append("END:VCALENDAR")
    return "\r\n".join(satirlar)


# ============== PROGRAM -> EXCEL (.xlsx) DÖNÜŞÜMÜ ==============
def programdan_excel_json_cikar(client_gemini, program_metni):
    talimat = (
        "Aşağıdaki metinde bir antrenman/beslenme/aktivite programı var. "
        "Bunu SADECE aşağıdaki JSON formatında, başka hiçbir açıklama/metin "
        "eklemeden döndür:\n\n"
        '[{"gun": "Pazartesi", "kategori": "Antrenman", '
        '"baslik": "Bacak Antrenmanı", '
        '"detay": "Squat 4x8, Leg Press 3x12, ...", '
        '"sure_kalori": "60 dk"}, ...]\n\n'
        "Kurallar:\n"
        "- kategori değeri 'Antrenman', 'Beslenme' ya da 'Diğer' olmalı\n"
        "- detay alanına set/tekrar sayılarını, öğün içeriğini vb. eksiksiz yaz\n"
        "- Metinde gerçekten bir program yoksa boş liste [] döndür\n\n"
        f"METİN:\n{program_metni}"
    )
    for model_adi in GEMINI_MODEL_LISTESI:
        try:
            yanit = client_gemini.models.generate_content(model=model_adi, contents=talimat)
            metin = yanit.text.strip()
            metin = re.sub(r"^```json\s*|\s*```$", "", metin.strip(), flags=re.MULTILINE)
            metin = metin.strip("`").strip()
            return json.loads(metin)
        except Exception:
            continue
    return None


def excel_plan_olustur(satirlar, baslik="Program"):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = baslik[:31] if baslik else "Program"

    basliklar = ["Gün", "Kategori", "Başlık", "Detay", "Süre/Kalori"]
    baslik_yazi_tipi = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    baslik_dolgu = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    hucre_yazi_tipi = Font(name="Arial", size=10)
    kenarlik = Border(*(Side(style="thin", color="D1D5DB"),) * 4)

    for sutun, metin in enumerate(basliklar, start=1):
        hucre = ws.cell(row=1, column=sutun, value=metin)
        hucre.font = baslik_yazi_tipi
        hucre.fill = baslik_dolgu
        hucre.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        hucre.border = kenarlik

    for satir_no, e in enumerate(satirlar, start=2):
        degerler = [e.get("gun", ""), e.get("kategori", ""), e.get("baslik", ""),
                    e.get("detay", ""), e.get("sure_kalori", "")]
        for sutun, deger in enumerate(degerler, start=1):
            hucre = ws.cell(row=satir_no, column=sutun, value=deger)
            hucre.font = hucre_yazi_tipi
            hucre.alignment = Alignment(vertical="top", wrap_text=True)
            hucre.border = kenarlik

    genislikler = [12, 12, 22, 45, 14]
    for i, genislik in enumerate(genislikler, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = genislik
    ws.freeze_panes = "A2"

    arabellek = BytesIO()
    wb.save(arabellek)
    return arabellek.getvalue()


# ============== KİMLİK DOĞRULAMA / KALICI SOHBET GEÇMİŞİ (PostgreSQL) ==============
DATABASE_URL = os.environ.get("DATABASE_URL", "")


def _postgres_semasini_hazirla():
    """Chainlit'in resmi tablo yapısını (users, threads, steps, elements,
    feedbacks) PostgreSQL'de bir kereliğine oluşturur. Tablolar zaten
    varsa (IF NOT EXISTS) hiçbir şey yapmaz, güvenle her açılışta çağrılabilir."""
    if not DATABASE_URL:
        return
    try:
        import psycopg2
        baglanti = psycopg2.connect(DATABASE_URL)
        baglanti.autocommit = True
        imlec = baglanti.cursor()
        imlec.execute("""
            CREATE TABLE IF NOT EXISTS users (
                "id" UUID PRIMARY KEY,
                "identifier" TEXT NOT NULL UNIQUE,
                "metadata" JSONB NOT NULL,
                "createdAt" TEXT
            );
            CREATE TABLE IF NOT EXISTS threads (
                "id" UUID PRIMARY KEY,
                "createdAt" TEXT,
                "name" TEXT,
                "userId" UUID,
                "userIdentifier" TEXT,
                "tags" TEXT[],
                "metadata" JSONB,
                FOREIGN KEY ("userId") REFERENCES users("id") ON DELETE CASCADE
            );
            CREATE TABLE IF NOT EXISTS steps (
                "id" UUID PRIMARY KEY,
                "name" TEXT NOT NULL,
                "type" TEXT NOT NULL,
                "threadId" UUID NOT NULL,
                "parentId" UUID,
                "streaming" BOOLEAN NOT NULL,
                "waitForAnswer" BOOLEAN,
                "isError" BOOLEAN,
                "metadata" JSONB,
                "tags" TEXT[],
                "input" TEXT,
                "output" TEXT,
                "createdAt" TEXT,
                "command" TEXT,
                "start" TEXT,
                "end" TEXT,
                "generation" JSONB,
                "showInput" TEXT,
                "language" TEXT,
                "indent" INT,
                "defaultOpen" BOOLEAN,
                "modes" JSONB,
                FOREIGN KEY ("threadId") REFERENCES threads("id") ON DELETE CASCADE
            );
            CREATE TABLE IF NOT EXISTS elements (
                "id" UUID PRIMARY KEY,
                "threadId" UUID,
                "type" TEXT,
                "url" TEXT,
                "chainlitKey" TEXT,
                "name" TEXT NOT NULL,
                "display" TEXT,
                "objectKey" TEXT,
                "size" TEXT,
                "page" INT,
                "language" TEXT,
                "forId" UUID,
                "mime" TEXT,
                "props" JSONB,
                FOREIGN KEY ("threadId") REFERENCES threads("id") ON DELETE CASCADE
            );
            CREATE TABLE IF NOT EXISTS feedbacks (
                "id" UUID PRIMARY KEY,
                "forId" UUID NOT NULL,
                "threadId" UUID NOT NULL,
                "value" INT NOT NULL,
                "comment" TEXT,
                FOREIGN KEY ("threadId") REFERENCES threads("id") ON DELETE CASCADE
            );

            -- Chainlit sürümü ilerledikçe 'steps' tablosuna yeni sütunlar
            -- ekleniyor. Kurulu sürümle uyumlu olmak için, bilinen tüm
            -- olası eksik sütunları burada güvenle (varsa dokunmadan) ekliyoruz.
            ALTER TABLE steps ADD COLUMN IF NOT EXISTS "command" TEXT;
            ALTER TABLE steps ADD COLUMN IF NOT EXISTS "defaultOpen" BOOLEAN;
            ALTER TABLE steps ADD COLUMN IF NOT EXISTS "modes" JSONB;
            ALTER TABLE steps ADD COLUMN IF NOT EXISTS "autoCollapse" BOOLEAN;
        """)
        imlec.close()
        baglanti.close()
        print("PostgreSQL şeması hazır (tablolar mevcut ya da oluşturuldu).")
    except Exception as e:
        print(f"PostgreSQL şeması hazırlanırken hata: {e}")


_postgres_semasini_hazirla()


@cl.password_auth_callback
def auth_callback(username: str, password: str):
    # Sadece kendi kullanımın için - herhangi bir kullanıcı adıyla giriş
    # kabul edilir, gerçek bir güvenlik katmanı değil.
    return cl.User(identifier=username or "ben")


@cl.data_layer
def get_data_layer():
    _conninfo = DATABASE_URL.replace("postgresql://", "postgresql+asyncpg://", 1)
    return SQLAlchemyDataLayer(conninfo=_conninfo)


# ============== CHAINLIT OLAYLARI ==============
@cl.on_chat_start
async def basla():
    istemcileri_al()
    cl.user_session.set("gecmis", [])
    await cl.Message(
        content=f"Merhaba! Ben **{UYGULAMA_ADI}**, senin kişisel hybrid antrenörünüm. "
                f"Bugün nasıl yardımcı olabilirim?\n\n"
                f"💡 İpucu: Bir fotoğraf ya da ses kaydı eklemek için mesaj kutusundaki "
                f"ataç ikonunu kullanabilirsin. Eski sohbet dosyalarını (.json) da aynı "
                f"şekilde ekleyip arşive dahil edebilirsin."
    ).send()


@cl.on_chat_resume
async def sohbete_devam(thread):
    """Sol menüden eski bir sohbete tıklandığında çalışır — hem mesaj
    kutusunu aktif hale getirir hem de Koçum'un o konuşmanın bağlamını
    (gecmis) hatırlamasını sağlar."""
    istemcileri_al()
    gecmis = []
    for adim in thread.get("steps", []):
        tur = adim.get("type", "")
        if tur == "user_message":
            metin = adim.get("output") or adim.get("input") or ""
            if metin:
                gecmis.append({"role": "user", "parts": [{"text": metin}]})
        elif tur == "assistant_message":
            metin = adim.get("output") or ""
            if metin:
                gecmis.append({"role": "model", "parts": [{"text": metin}]})
    cl.user_session.set("gecmis", gecmis)


def _sohbet_json_arsivle(koleksiyon, dosya_yolu, dosya_adi):
    """Eski Streamlit sohbet JSON dosyasını okuyup arşive (ChromaDB) ekler."""
    with open(dosya_yolu, 'r', encoding='utf-8') as f:
        veri = json.load(f)

    mesajlar = veri.get("mesajlar", [])
    if not mesajlar:
        return 0

    baslik = veri.get("baslik", dosya_adi)
    satirlar = [f"# Eski Sohbet: {baslik}\n"]
    for m in mesajlar:
        rol = "Kullanıcı" if m.get("role") == "user" else "Koçum"
        icerik = m.get("content", "")
        if icerik:
            satirlar.append(f"{rol}: {icerik}")
    metin = "\n\n".join(satirlar)

    if len(metin.strip()) < 50:
        return 0

    boyut, ortusme = 800, 150
    parcalar, baslangic = [], 0
    while baslangic < len(metin):
        parcalar.append(metin[baslangic:baslangic + boyut])
        baslangic += (boyut - ortusme)

    etiket = f"eski_sohbet_{uuid.uuid4().hex[:8]}"
    ids = [f"{etiket}_parca_{j}" for j in range(len(parcalar))]
    metadatalar = [{"video_id": f"Eski sohbet: {baslik}", "kaynak": "eski_sohbet"} for _ in parcalar]
    koleksiyon.add(documents=parcalar, ids=ids, metadatas=metadatalar)
    return len(parcalar)


async def _eski_sohbeti_gercek_thread_yap(dosya_yolu, dosya_adi, kullanici_identifier):
    """Eski sohbet JSON'unu, Postgres'te GERÇEK bir thread + steps olarak
    oluşturur. Böylece sol menüde tıklanabilir, devam edilebilir bir
    sohbet olarak görünür (sadece arka plan bilgisi değil)."""
    if not DATABASE_URL or not kullanici_identifier:
        return None

    import asyncpg

    with open(dosya_yolu, 'r', encoding='utf-8') as f:
        veri = json.load(f)
    mesajlar = veri.get("mesajlar", [])
    if not mesajlar:
        return None

    baslik = veri.get("baslik", dosya_adi)

    conn = await asyncpg.connect(DATABASE_URL)
    try:
        kullanici_satiri = await conn.fetchrow(
            'SELECT id FROM users WHERE identifier = $1', kullanici_identifier
        )
        if not kullanici_satiri:
            return None
        kullanici_id = kullanici_satiri['id']

        thread_id = str(uuid.uuid4())
        simdi = datetime.utcnow().isoformat() + "Z"
        await conn.execute(
            'INSERT INTO threads (id, "createdAt", name, "userId", "userIdentifier", tags, metadata) '
            'VALUES ($1, $2, $3, $4, $5, $6, $7)',
            thread_id, simdi, f"📜 {baslik}", kullanici_id, kullanici_identifier, [], '{}'
        )

        for m in mesajlar:
            icerik = m.get("content", "")
            if not icerik:
                continue
            rol = m.get("role")
            tur = "user_message" if rol == "user" else "assistant_message"
            ad = kullanici_identifier if tur == "user_message" else UYGULAMA_ADI
            zaman = datetime.utcnow().isoformat() + "Z"
            adim_id = str(uuid.uuid4())
            await conn.execute(
                'INSERT INTO steps (id, name, type, "threadId", "parentId", streaming, input, '
                '"isError", output, "createdAt", start, "end", "defaultOpen", "autoCollapse", '
                '"showInput", metadata) '
                'VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10,$11,$12,$13,$14,$15,$16)',
                adim_id, ad, tur, thread_id, None, False, icerik, False, icerik,
                zaman, zaman, zaman, False, False, "json", '{}'
            )
        return thread_id
    finally:
        await conn.close()


def _pcm_wav_yap(pcm_bytes, ornek_hizi=24000, kanal=1, ornek_genisligi=2):
    """Chainlit'ten gelen ham (headersiz) PCM ses verisini, Gemini'nin
    anlayabileceği düzgün bir WAV dosyasına çevirir."""
    import wave
    tampon = BytesIO()
    with wave.open(tampon, 'wb') as wf:
        wf.setnchannels(kanal)
        wf.setsampwidth(ornek_genisligi)
        wf.setframerate(ornek_hizi)
        wf.writeframes(pcm_bytes)
    return tampon.getvalue()


@cl.on_audio_start
async def ses_baslar():
    cl.user_session.set("ses_parcalari", [])
    return True


@cl.on_audio_chunk
async def ses_parcasi_geldi(chunk: cl.InputAudioChunk):
    parcalar = cl.user_session.get("ses_parcalari") or []
    parcalar.append(chunk.data)
    cl.user_session.set("ses_parcalari", parcalar)


@cl.on_audio_end
async def ses_biter():
    client_gemini, koleksiyon = istemcileri_al()
    parcalar = cl.user_session.get("ses_parcalari") or []
    cl.user_session.set("ses_parcalari", [])

    if not parcalar:
        return

    pcm_veri = b"".join(parcalar)
    wav_veri = _pcm_wav_yap(pcm_veri)

    async with cl.Step(name="Ses yazıya çevriliyor", type="tool"):
        yazi = await cl.make_async(ses_yaziya_cevir)(client_gemini, wav_veri, "audio/wav")

    if not yazi:
        await cl.Message(content="Ses anlaşılamadı, tekrar dener misin? "
                                  "(Sorun devam ederse ataç ikonuyla ses dosyası da ekleyebilirsin.)").send()
        return

    kullanici_mesaji = cl.Message(content=yazi, author="Kullanıcı")
    await kullanici_mesaji.send()
    await mesaj_geldi(kullanici_mesaji)


@cl.on_message
async def mesaj_geldi(message: cl.Message):
    client_gemini, koleksiyon = istemcileri_al()
    soru = message.content or ""

    gorsel_b64, gorsel_mime = None, None
    json_sonuclari = []
    for el in message.elements:
        mime = getattr(el, "mime", "") or ""
        yol = getattr(el, "path", None)
        ad = getattr(el, "name", "") or ""
        if not yol:
            continue
        if mime.startswith("image"):
            with open(yol, "rb") as f:
                gorsel_b64 = base64.b64encode(f.read()).decode("utf-8")
            gorsel_mime = mime
        elif mime.startswith("audio"):
            with open(yol, "rb") as f:
                ses_bytes = f.read()
            async with cl.Step(name="Ses yazıya çevriliyor", type="tool"):
                yazi = await cl.make_async(ses_yaziya_cevir)(client_gemini, ses_bytes, mime)
            if yazi:
                soru = (soru + "\n" + yazi).strip() if soru else yazi
        elif ad.endswith(".json") or mime == "application/json":
            try:
                sayi = await cl.make_async(_sohbet_json_arsivle)(koleksiyon, yol, ad)
                json_sonuclari.append((ad, sayi))
            except Exception as e:
                json_sonuclari.append((ad, f"HATA: {e}"))
                continue

            try:
                kullanici = cl.user_session.get("user")
                kullanici_id = kullanici.identifier if kullanici else None
                thread_id = await _eski_sohbeti_gercek_thread_yap(yol, ad, kullanici_id)
                if thread_id:
                    json_sonuclari.append((ad, "🧵 Gerçek sohbet olarak da eklendi — sol menüde "
                                                "göreceksin, tıklayıp devam edebilirsin"))
            except Exception as e:
                json_sonuclari.append((ad, f"(gerçek sohbet oluşturulamadı: {e})"))

    if json_sonuclari:
        satirlar = ["📥 **Eski sohbet(ler) arşive eklendi:**"]
        for ad, sonuc in json_sonuclari:
            if isinstance(sonuc, int):
                satirlar.append(f"- {ad}: {sonuc} parça eklendi" if sonuc else f"- {ad}: boş, atlandı")
            else:
                satirlar.append(f"- {ad}: {sonuc}")
        await cl.Message(content="\n".join(satirlar)).send()
        if not soru:
            return

    if not soru:
        await cl.Message(content="Bir metin, fotoğraf ya da ses kaydı gönderir misin?").send()
        return

    async with cl.Step(name="Arşiv taranıyor", type="tool"):
        bulunan = await cl.make_async(koleksiyon.query)(
            query_texts=[soru], n_results=KAC_PARCA_GETIRILSIN
        )
        baglam, kaynaklar = "", []
        if bulunan['documents'][0]:
            baglam, kaynaklar = baglami_hazirla(bulunan)

    gecmis = cl.user_session.get("gecmis", [])
    cevap = await cl.make_async(cevap_uret)(client_gemini, soru, baglam, gecmis, gorsel_b64, gorsel_mime)

    gecmis = gecmis + [
        {"role": "user", "parts": [{"text": soru}]},
        {"role": "model", "parts": [{"text": cevap}]},
    ]
    cl.user_session.set("gecmis", gecmis)

    actions = [
        cl.Action(name="takvim_yap", payload={"metin": cevap}, label="📅 Takvime Hazırla"),
        cl.Action(name="excel_yap", payload={"metin": cevap}, label="📊 Excel Yap"),
    ]

    nihai_metin = cevap
    if kaynaklar:
        kaynak_metni = "\n".join(
            f"- {k['link']}" if k.get('link') else f"- 📜 {k.get('baslik', 'Eski sohbet')}"
            for k in kaynaklar[:10]
        )
        nihai_metin = f"{cevap}\n\n---\n**🔍 Kullanılan kaynaklar:**\n{kaynak_metni}"

    await cl.Message(content=nihai_metin, actions=actions).send()


@cl.action_callback("takvim_yap")
async def takvim_yap(action: cl.Action):
    metin = action.payload.get("metin", "")
    client_gemini, _ = istemcileri_al()
    async with cl.Step(name="Program takvime çevriliyor", type="tool"):
        etkinlikler = await cl.make_async(programdan_json_cikar)(client_gemini, metin)
    if not etkinlikler:
        await cl.Message(content="Bu mesajda takvime çevrilecek bir program bulunamadı.").send()
        return
    ics_veri = ics_plan_olustur(etkinlikler, baslik_oneki=UYGULAMA_ADI)
    dosya = cl.File(name="program.ics", content=ics_veri.encode("utf-8"))
    await cl.Message(content="📅 Takvim dosyan hazır, indirip Google Takvim'e aktarabilirsin:",
                      elements=[dosya]).send()


@cl.action_callback("excel_yap")
async def excel_yap(action: cl.Action):
    metin = action.payload.get("metin", "")
    client_gemini, _ = istemcileri_al()
    async with cl.Step(name="Program Excel'e çevriliyor", type="tool"):
        satirlar = await cl.make_async(programdan_excel_json_cikar)(client_gemini, metin)
    if not satirlar:
        await cl.Message(content="Bu mesajda Excel'e çevrilecek bir program bulunamadı.").send()
        return
    excel_veri = excel_plan_olustur(satirlar, baslik=UYGULAMA_ADI)
    dosya = cl.File(name="program.xlsx", content=excel_veri)
    await cl.Message(content="📊 Excel dosyan hazır:", elements=[dosya]).send()
